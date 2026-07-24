from __future__ import annotations

import hashlib
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output/rm2_sentiment/validation/human_master_v4"
MASTER = OUT_DIR / "human_annotation_master_v4.csv"
CHECKSUMS = OUT_DIR / "source_annotation_checksums.csv"
LEAKAGE = OUT_DIR / "master_split_leakage_audit.csv"
COMMENT_CONFLICTS = OUT_DIR / "master_comment_id_conflicts.csv"
TEXT_CONFLICTS = OUT_DIR / "master_text_cluster_conflicts.csv"

ANNOTATOR_WORKBOOKS = [
    OUT_DIR / "sentiment_v4_development_annotator_1.xlsx",
    OUT_DIR / "sentiment_v4_development_annotator_2.xlsx",
    OUT_DIR / "sentiment_v4_locked_test_annotator_1.xlsx",
    OUT_DIR / "sentiment_v4_locked_test_annotator_2.xlsx",
]

LABELS = {"Negative", "Neutral", "Positive", "Uncertain", "No Text"}
PENDING_ROLES = {"DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"}
DEV_ROLES = {"HISTORICAL_DEVELOPMENT_FINAL", "DEVELOPMENT_NEW_PENDING", "DEVELOPMENT_NEW_FINAL"}
LOCKED_ROLES = {"LOCKED_TEST_NEW_PENDING", "LOCKED_TEST_NEW_FINAL"}
ROLE_VALUES = {
    "HISTORICAL_DEVELOPMENT_FINAL",
    "DEVELOPMENT_NEW_PENDING",
    "DEVELOPMENT_NEW_FINAL",
    "LEGACY_TEST_PROVENANCE",
    "LOCKED_TEST_NEW_PENDING",
    "LOCKED_TEST_NEW_FINAL",
    "EXCLUDED",
}
ANNOTATOR_COLUMNS = {
    "annotation_id",
    "comment_id",
    "video_id",
    "product_category",
    "brand_or_video_context",
    "comment_text",
    "human_label",
    "confidence_annotation",
    "needs_context",
    "annotator_notes",
}
BANNED_TOKENS = ["prediction", "probability", "threshold", "hcc", "actor", "goal", "model", "disagreement"]


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def is_inj(value: object) -> bool:
    text = "" if pd.isna(value) else str(value).strip().upper()
    return text.startswith("INJ") or "SYNTHETIC" in text


class SentimentMasterAnnotationV4Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.master = read_csv(MASTER)

    def test_old_source_files_unchanged_since_inventory(self):
        checksums = read_csv(CHECKSUMS)
        for _, row in checksums.iterrows():
            path = ROOT / row["path"]
            if path.exists():
                self.assertEqual(sha256_file(path), row["sha256"], row["path"])

    def test_comment_id_unique(self):
        self.assertEqual(int(self.master["comment_id"].duplicated().sum()), 0)

    def test_roles_are_locked_to_vocabulary(self):
        self.assertTrue(set(self.master["annotation_role"]).issubset(ROLE_VALUES))
        pending = self.master.loc[self.master["annotation_role"].isin(PENDING_ROLES)]
        self.assertTrue(pending["split_lock"].isin({"DEVELOPMENT_PENDING_LOCKED", "LOCKED_TEST_PENDING_LOCKED"}).all())

    def test_no_inj_in_pending_annotation_rows(self):
        pending = self.master.loc[self.master["annotation_role"].isin(PENDING_ROLES)]
        self.assertEqual(int(pending["comment_id"].map(is_inj).sum()), 0)

    def test_hard_duplicate_leakage_zero(self):
        leakage = read_csv(LEAKAGE)
        hard = leakage.loc[leakage["audit_type"].isin(["comment_id_development_locked_overlap", "text_cluster_development_locked_overlap"])]
        self.assertTrue(hard["status"].eq("PASS").all())
        self.assertTrue(hard["observed"].astype(int).eq(0).all())

    def test_conflict_audits_pass(self):
        self.assertTrue(read_csv(COMMENT_CONFLICTS)["status"].eq("PASS").all())
        self.assertTrue(read_csv(TEXT_CONFLICTS)["status"].eq("PASS").all())

    def test_legacy_test_not_in_development(self):
        legacy_dev = self.master.loc[
            self.master["annotation_role"].isin(DEV_ROLES)
            & self.master["source_file"].str.contains("locked_test|final_test", case=False, regex=True, na=False)
        ]
        self.assertEqual(len(legacy_dev), 0)

    def test_new_locked_test_not_in_development(self):
        dev_ids = set(self.master.loc[self.master["annotation_role"].isin(DEV_ROLES), "comment_id"])
        locked_ids = set(self.master.loc[self.master["annotation_role"].isin(LOCKED_ROLES), "comment_id"])
        self.assertFalse(dev_ids & locked_ids)

    def test_annotator_workbooks_do_not_expose_predictions(self):
        for path in ANNOTATOR_WORKBOOKS:
            wb = load_workbook(path, read_only=True, data_only=True)
            data_sheets = [ws for ws in wb.worksheets if ws.title.upper() != "README"]
            self.assertEqual(len(data_sheets), 1, path.name)
            headers = [str(value) for value in next(data_sheets[0].iter_rows(min_row=1, max_row=1, values_only=True))]
            self.assertTrue(ANNOTATOR_COLUMNS.issubset(headers), path.name)
            lower_headers = [header.lower() for header in headers]
            for token in BANNED_TOKENS:
                allowed = {"annotation_id", "confidence_annotation", "annotator_notes"}
                self.assertFalse(any(token in header and header not in allowed for header in lower_headers), path.name)

    def test_label_vocabulary(self):
        non_blank = set(self.master["final_human_label"]) - {""}
        self.assertTrue(non_blank.issubset(LABELS))

    def test_disagreement_not_auto_filled_for_pending_rows(self):
        pending = self.master.loc[self.master["annotation_role"].isin(PENDING_ROLES)]
        self.assertTrue(pending["annotator_1_label"].eq("").all())
        self.assertTrue(pending["annotator_2_label"].eq("").all())
        self.assertTrue(pending["adjudicated_label"].eq("").all())
        self.assertTrue(pending["final_human_label"].eq("").all())

    def test_final_label_only_human_sources(self):
        historical = self.master.loc[self.master["annotation_role"].eq("HISTORICAL_DEVELOPMENT_FINAL")]
        self.assertTrue(historical["final_human_label"].isin(LABELS).all())
        pending = self.master.loc[self.master["annotation_role"].isin(PENDING_ROLES)]
        self.assertTrue(pending["final_human_label"].eq("").all())

    def test_pending_counts(self):
        counts = self.master["annotation_role"].value_counts()
        self.assertEqual(int(counts.get("DEVELOPMENT_NEW_PENDING", 0)), 1300)
        self.assertEqual(int(counts.get("LOCKED_TEST_NEW_PENDING", 0)), 700)
        self.assertEqual(int(self.master["annotation_role"].isin(PENDING_ROLES).sum()), 2000)


if __name__ == "__main__":
    unittest.main()
