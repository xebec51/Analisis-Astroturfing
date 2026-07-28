from __future__ import annotations

import json
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from scripts.train_rm2_sentiment_indobert_v5_development import model_input


ROOT = Path(__file__).resolve().parents[1]
FAIR_DIR = ROOT / "output/rm2_sentiment/experiments/fair_same_test_comparison"
V5_DEV_DIR = ROOT / "output/rm2_sentiment/validation/human_v5"
V5_LOCK_DIR = ROOT / "output/rm2_sentiment/validation/human_v5_locked_test"
V5_EXP_DIR = ROOT / "output/rm2_sentiment/experiments/indobert_v5_development"
SENS_DIR = ROOT / "output/rm2_sentiment/sensitivity/indobert_v4"
V4_DECISION = ROOT / "output/rm2_sentiment/experiments/indobert_v4_final/locked_test_evaluation/FINAL_ACCEPTANCE_DECISION.json"
CANONICAL_MODEL = ROOT / "output/rm2_sentiment/final/CANONICAL_MODEL.json"

LABELS = ["Negative", "Neutral", "Positive"]
FORBIDDEN_ANNOTATOR_COLUMNS = {
    "v2_probability_negative",
    "v2_probability_neutral",
    "v2_probability_positive",
    "v2_confidence",
    "final_sentiment_label",
    "sampling_stratum",
}


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)


def norm_text(value: object) -> str:
    return " ".join(str(value).lower().split())


class RM2SentimentFairComparisonV5Tests(unittest.TestCase):
    def test_v4_strict_decision_is_preserved(self):
        decision = read_json(V4_DECISION)
        self.assertEqual(decision["status"], "INDOBERT_V4_NOT_ACCEPTED_KEEP_V2")
        self.assertFalse(decision["accepted"])
        self.assertFalse(CANONICAL_MODEL.exists())

    def test_same_test_predictions_use_identical_denominator(self):
        predictions = read_csv(FAIR_DIR / "same_test_predictions.csv")
        self.assertEqual(len(predictions), 672)
        self.assertEqual(predictions["comment_id"].nunique(), 672)
        self.assertEqual(predictions["final_human_label"].value_counts().reindex(LABELS).fillna(0).astype(int).to_dict(), {
            "Negative": 160,
            "Neutral": 294,
            "Positive": 218,
        })
        self.assertTrue(predictions["v4_argmax_label"].isin(LABELS).all())
        self.assertTrue(predictions["v2_forced_label"].isin(LABELS).all())

    def test_same_test_metrics_and_abstention_accounting(self):
        metrics = read_csv(FAIR_DIR / "same_test_metrics.csv")
        modes = set(metrics["mode"])
        self.assertEqual(modes, {"V2_NATIVE_POLICY", "V2_FORCED_THREE_CLASS", "V4_ARGMAX_THREE_CLASS"})
        native = metrics.loc[metrics["mode"].eq("V2_NATIVE_POLICY")].iloc[0]
        forced = metrics.loc[metrics["mode"].eq("V2_FORCED_THREE_CLASS")].iloc[0]
        v4 = metrics.loc[metrics["mode"].eq("V4_ARGMAX_THREE_CLASS")].iloc[0]
        self.assertEqual(int(float(native["number_abstained"])), 189)
        self.assertAlmostEqual(float(native["coverage"]), 483 / 672)
        self.assertEqual(int(float(forced["number_abstained"])), 0)
        self.assertEqual(int(float(v4["number_abstained"])), 0)
        self.assertGreater(float(v4["macro_f1"]), float(forced["macro_f1"]))
        self.assertGreater(float(v4["full_set_accuracy_abstain_wrong"]), float(forced["full_set_accuracy_abstain_wrong"]))

    def test_same_test_paired_outputs_exist(self):
        for name in [
            "same_test_mcnemar.json",
            "same_test_bootstrap_ci.json",
            "same_test_disagreement.csv",
            "same_test_confusion_matrices.json",
            "SAME_TEST_COMPARISON_MANIFEST.json",
        ]:
            self.assertTrue((FAIR_DIR / name).exists(), name)
        manifest = read_json(FAIR_DIR / "SAME_TEST_COMPARISON_MANIFEST.json")
        self.assertEqual(manifest["status"], "INDOBERT_V4_SAME_TEST_FAIR_COMPARISON")
        self.assertFalse(manifest["locked_test_used_for_tuning"])
        self.assertFalse(manifest["canonical_model_changed"])

    def test_v5_annotation_packages_have_no_overlap_or_auto_labels(self):
        dev = read_csv(V5_DEV_DIR / "sentiment_v5_development_candidates.csv")
        locked = read_csv(V5_LOCK_DIR / "sentiment_v5_locked_test_candidates.csv")
        self.assertEqual(len(dev), 1000)
        self.assertEqual(len(locked), 700)
        self.assertFalse(set(dev["comment_id"]) & set(locked["comment_id"]))
        self.assertFalse((set(dev["comment_text"].map(norm_text)) & set(locked["comment_text"].map(norm_text))) - {""})
        for frame in [dev, locked]:
            self.assertFalse(frame["final_human_label"].isin(LABELS).any())
            self.assertFalse(frame["sentiment_toward_target"].isin(LABELS).any())
            self.assertFalse(frame["comment_id"].astype(str).str.contains("INJ", case=False, regex=False).any())

    def test_annotator_workbooks_do_not_expose_sampling_or_predictions(self):
        for path in [
            V5_DEV_DIR / "sentiment_v5_development_annotator_1.xlsx",
            V5_DEV_DIR / "sentiment_v5_development_annotator_2.xlsx",
            V5_LOCK_DIR / "sentiment_v5_locked_test_annotator_1.xlsx",
            V5_LOCK_DIR / "sentiment_v5_locked_test_annotator_2.xlsx",
        ]:
            wb = load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = {cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))}
            self.assertFalse(headers & FORBIDDEN_ANNOTATOR_COLUMNS, path)
            self.assertIn("sentiment_toward_target", headers)
            self.assertIn("sentiment_overall", headers)

    def test_v5_acceptance_config_is_preregistered_before_locked_test_open(self):
        config = read_json(ROOT / "configs/rm2_sentiment_v5_acceptance_preregistered.json")
        self.assertEqual(config["status"], "RM2_SENTIMENT_V5_ACCEPTANCE_PREREGISTERED_BEFORE_LOCKED_TEST_V5_OPENED")
        self.assertEqual(config["primary_label"], "sentiment_toward_target")
        self.assertTrue(config["coverage_policy"]["do_not_compare_v2_covered_accuracy_to_v5_full_coverage_accuracy_as_gate"])
        locked_manifest = read_json(V5_LOCK_DIR / "LOCKED_TEST_V5_FREEZE_MANIFEST.json")
        self.assertIn(
            locked_manifest["status"],
            {
                "LOCKED_TEST_V5_CANDIDATE_LIST_FROZEN_PENDING_HUMAN_LABELS",
                "LOCKED_TEST_V5_PENDING_HUMAN_ADJUDICATION",
                "LOCKED_TEST_V5_FINAL_LABELS_FROZEN_SEALED",
            },
        )
        self.assertIn(
            locked_manifest["final_label_freeze_status"],
            {
                "BLOCKED_PENDING_TWO_ANNOTATORS_AND_ADJUDICATION",
                "BLOCKED_PENDING_ADJUDICATION",
                "FROZEN_NOT_OPEN_FOR_MODEL_SELECTION",
            },
        )
        self.assertFalse(locked_manifest.get("locked_test_v5_used_for_training_or_selection", False))

    def test_v5_development_pipeline_plan_is_development_only(self):
        manifest = read_json(V5_EXP_DIR / "INDOBERT_V5_DEVELOPMENT_PIPELINE_MANIFEST.json")
        self.assertIn(
            manifest["status"],
            {
                "INDOBERT_V5_DEVELOPMENT_PIPELINE_READY_PENDING_HUMAN_LABELS",
                "INDOBERT_V5_DEVELOPMENT_PIPELINE_READY_FINAL_HUMAN_LABELS",
                "INDOBERT_V5_DEVELOPMENT_FOLDS_READY",
            },
        )
        self.assertEqual(manifest["candidate_trial_count"], 810)
        self.assertFalse(manifest["locked_test_v4_errors_used"])
        self.assertFalse(manifest["locked_test_v5_labels_used_for_training_or_selection"])
        self.assertEqual(manifest["validation"].get("locked_labels_available_for_training", 0), 0)
        grid = read_csv(V5_EXP_DIR / "candidate_grid_manifest.csv")
        self.assertEqual(len(grid), 810)
        if manifest["status"] == "INDOBERT_V5_DEVELOPMENT_FOLDS_READY":
            folds = read_csv(V5_EXP_DIR / "development_grouped_fold_assignments.csv")
            locked_candidates = read_csv(V5_LOCK_DIR / "sentiment_v5_locked_test_candidates.csv")
            self.assertEqual(len(folds), 2931)
            self.assertEqual(set(folds["seed"].astype(int)), {42, 52, 62})
            self.assertFalse(folds["comment_id"].isin(locked_candidates["comment_id"]).any())

    def test_target_context_parent_preprocessing(self):
        row = pd.Series({
            "brand_or_video_context": "Azarine",
            "parent_comment_text": "aman untuk pemula?",
            "comment_text": "iya cocok di aku",
        })
        text = model_input(row, "target_context_parent_comment")
        self.assertEqual(text, "Azarine [SEP] aman untuk pemula? [SEP] iya cocok di aku")

    def test_sensitivity_outputs_are_noncanonical(self):
        manifest = read_json(SENS_DIR / "INDOBERT_V4_SENSITIVITY_MANIFEST.json")
        self.assertEqual(manifest["status"], "EXPLORATORY_SENSITIVITY_ANALYSIS_NOT_CANONICAL")
        self.assertFalse(manifest["canonical_model_changed"])
        self.assertEqual(manifest["input_rows"], 33063)
        predictions = read_csv(SENS_DIR / "indobert_v4_comment_sentiment_sensitivity.csv")
        self.assertEqual(len(predictions), 33063)
        self.assertEqual(predictions["comment_id"].nunique(), 33063)
        self.assertTrue(predictions["sentiment_label"].isin([*LABELS, "No Text"]).all())
        goals = read_csv(SENS_DIR / "tables/v2_vs_v4_hcc_goal_changes.csv")
        self.assertEqual(int(goals["goal_changed"].astype(str).str.lower().eq("true").sum()), 3)

    def test_completed_v5_annotations_imported_and_finalized(self):
        manifest = read_json(V5_DEV_DIR / "SENTIMENT_V5_IMPORT_MANIFEST.json")
        self.assertIn(
            manifest["status"],
            {
                "SENTIMENT_V5_PENDING_HUMAN_ADJUDICATION",
                "SENTIMENT_V5_FINAL_HUMAN_LABELS_FROZEN",
                "SENTIMENT_V5_FINAL_HUMAN_LABELS_IMPORTED",
            },
        )
        self.assertEqual(manifest["development"]["rows"], 1000)
        self.assertEqual(manifest["development"]["disagreement_rows"], 23)
        self.assertGreaterEqual(manifest["development"]["sentiment_toward_target"]["cohen_kappa"], 0.95)
        self.assertEqual(manifest["locked_test"]["rows"], 700)
        self.assertEqual(manifest["locked_test"]["disagreement_rows"], 21)
        self.assertGreaterEqual(manifest["locked_test"]["sentiment_toward_target"]["cohen_kappa"], 0.95)

        final_manifest = read_json(V5_DEV_DIR / "SENTIMENT_V5_FINAL_IMPORT_MANIFEST.json")
        self.assertEqual(final_manifest["status"], "SENTIMENT_V5_FINAL_HUMAN_LABELS_IMPORTED")
        self.assertFalse(final_manifest["methodology"]["model_predictions_used_for_labels"])
        self.assertFalse(final_manifest["methodology"]["auto_final_labels_for_disagreements"])
        self.assertTrue(final_manifest["methodology"]["final_labels_from_human_agreement_or_adjudication_only"])
        self.assertFalse(final_manifest["methodology"]["locked_test_v5_used_for_training_or_selection"])
        self.assertEqual(final_manifest["development"]["evaluable_three_class_rows"], 977)
        self.assertEqual(
            final_manifest["development"]["label_counts"],
            {"Negative": 178, "Neutral": 569, "Positive": 230, "Uncertain": 13, "No Text": 10},
        )
        self.assertEqual(final_manifest["locked_test_v5"]["evaluable_three_class_rows"], 687)
        self.assertEqual(
            final_manifest["locked_test_v5"]["label_counts"],
            {"Negative": 134, "Neutral": 380, "Positive": 173, "Uncertain": 9, "No Text": 4},
        )
        self.assertTrue(final_manifest["leakage_audit"]["hard_leakage_pass"])

        final_dev = read_csv(V5_DEV_DIR / "sentiment_v5_development_final_registry.csv")
        final_locked = read_csv(V5_LOCK_DIR / "sentiment_v5_locked_test_final_frozen.csv")
        self.assertEqual(len(final_dev), 1000)
        self.assertEqual(len(final_locked), 700)
        self.assertEqual(final_dev["comment_id"].nunique(), 1000)
        self.assertEqual(final_locked["comment_id"].nunique(), 700)
        self.assertFalse(set(final_dev["comment_id"]) & set(final_locked["comment_id"]))
        self.assertTrue(final_dev["final_human_label"].isin([*LABELS, "Uncertain", "No Text"]).all())
        self.assertTrue(final_locked["final_human_label"].isin([*LABELS, "Uncertain", "No Text"]).all())
        self.assertTrue(final_dev.loc[final_dev["adjudication_required"].astype(str).str.lower().eq("true"), "final_label_source"].eq("human_adjudication").all())
        self.assertTrue(final_locked.loc[final_locked["adjudication_required"].astype(str).str.lower().eq("true"), "final_label_source"].eq("human_adjudication").all())

        for path, expected_rows in [
            (V5_DEV_DIR / "sentiment_v5_development_adjudication.xlsx", 23),
            (V5_LOCK_DIR / "sentiment_v5_locked_test_adjudication.xlsx", 21),
        ]:
            workbook = load_workbook(path, read_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows())
            self.assertEqual(len(rows) - 1, expected_rows)
            headers = [cell.value for cell in rows[0]]
            adjudicated_idx = headers.index("adjudicated_label")
            labels = [row[adjudicated_idx].value for row in rows[1:]]
            self.assertEqual(sum(value not in {None, ""} for value in labels), expected_rows)
            self.assertTrue(all(value in {"Negative", "Neutral", "Positive", "Uncertain", "No Text"} for value in labels))


if __name__ == "__main__":
    unittest.main()
