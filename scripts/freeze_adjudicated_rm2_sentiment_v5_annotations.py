from __future__ import annotations

import hashlib
import json
import math
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEV_DIR = ROOT / "output/rm2_sentiment/validation/human_v5"
LOCK_DIR = ROOT / "output/rm2_sentiment/validation/human_v5_locked_test"

LABELS = ["Negative", "Neutral", "Positive", "Uncertain", "No Text"]
THREE_CLASS = ["Negative", "Neutral", "Positive"]

DEV_PRE_REGISTRY = DEV_DIR / "sentiment_v5_development_pre_adjudication_registry.csv"
LOCK_PRE_REGISTRY = LOCK_DIR / "sentiment_v5_locked_test_pre_adjudication_registry.csv"
DEV_ADJUDICATION = DEV_DIR / "sentiment_v5_development_adjudication.xlsx"
LOCK_ADJUDICATION = LOCK_DIR / "sentiment_v5_locked_test_adjudication.xlsx"
LOCK_ADJUDICATION_FALLBACK = DEV_DIR / "sentiment_v5_locked_test_adjudication.xlsx"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def rel(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return str(path)


def to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): to_jsonable(v) for k, v in value.items()}
    if isinstance(value, list | tuple):
        return [to_jsonable(v) for v in value]
    if isinstance(value, Path):
        return rel(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(to_jsonable(payload), indent=2), encoding="utf-8")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_dataframe(frame: pd.DataFrame, columns: list[str]) -> str:
    serialised = frame[columns].sort_values(columns).to_csv(index=False).encode("utf-8")
    return hashlib.sha256(serialised).hexdigest()


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)


def read_first_sheet(path: Path) -> pd.DataFrame:
    sheet = pd.ExcelFile(path).sheet_names[0]
    frame = pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
    frame.columns = [str(column).strip() for column in frame.columns]
    return frame


def normalize_label(value: object) -> str:
    text = "" if value is None or pd.isna(value) else str(value).strip()
    aliases = {
        "negative": "Negative",
        "neutral": "Neutral",
        "positive": "Positive",
        "uncertain": "Uncertain",
        "no text": "No Text",
        "no_text": "No Text",
        "notext": "No Text",
    }
    return aliases.get(text.lower(), text)


def parse_bool(value: object) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def normalize_text(value: object) -> str:
    text = "" if value is None or pd.isna(value) else str(value)
    text = unicodedata.normalize("NFKC", text)
    return " ".join(text.lower().split())


def label_counts(frame: pd.DataFrame, column: str = "final_human_label") -> dict[str, int]:
    values = frame[column].astype(str).map(normalize_label)
    return values.value_counts().reindex(LABELS, fill_value=0).astype(int).to_dict()


def pick_adjudication_workbook(canonical_path: Path, fallback_path: Path | None = None) -> Path:
    if canonical_path.exists():
        canonical = read_first_sheet(canonical_path)
        if "adjudicated_label" in canonical.columns:
            filled = canonical["adjudicated_label"].map(normalize_label).astype(str).str.strip().ne("").sum()
            if filled:
                return canonical_path
    if fallback_path and fallback_path.exists():
        fallback = read_first_sheet(fallback_path)
        if "adjudicated_label" in fallback.columns:
            filled = fallback["adjudicated_label"].map(normalize_label).astype(str).str.strip().ne("").sum()
            if filled:
                return fallback_path
    return canonical_path


def sync_adjudication_workbook(source_path: Path, target_path: Path) -> bool:
    if source_path == target_path:
        return False
    source = read_first_sheet(source_path)
    if "annotation_id" not in source.columns or "adjudicated_label" not in source.columns:
        raise ValueError(f"Missing adjudication columns in {source_path}")
    source_map = source.set_index("annotation_id", drop=False)

    workbook = load_workbook(target_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    annotation_col = headers.index("annotation_id") + 1
    label_col = headers.index("adjudicated_label") + 1
    note_col = headers.index("adjudication_note") + 1

    changed = False
    for row_idx in range(2, sheet.max_row + 1):
        annotation_id = str(sheet.cell(row=row_idx, column=annotation_col).value or "").strip()
        if annotation_id not in source_map.index:
            continue
        source_row = source_map.loc[annotation_id]
        label = normalize_label(source_row["adjudicated_label"])
        note = str(source_row.get("adjudication_note", "") or "").strip()
        if label:
            sheet.cell(row=row_idx, column=label_col).value = label
            sheet.cell(row=row_idx, column=note_col).value = note
            changed = True
    if changed:
        workbook.save(target_path)
    return changed


def validate_adjudication(registry: pd.DataFrame, adjudication: pd.DataFrame, split_name: str) -> pd.DataFrame:
    required = registry.loc[registry["adjudication_required"].map(parse_bool)].copy()
    if len(required) == 0:
        return pd.DataFrame(columns=["annotation_id", "adjudicated_label", "adjudication_note"])
    for column in ["annotation_id", "comment_id", "adjudicated_label"]:
        if column not in adjudication.columns:
            raise ValueError(f"{split_name} adjudication workbook missing column: {column}")
    adjudication = adjudication.copy()
    adjudication["adjudicated_label"] = adjudication["adjudicated_label"].map(normalize_label)
    adjudication["adjudication_note"] = adjudication.get("adjudication_note", "").astype(str)
    invalid = sorted(set(adjudication.loc[adjudication["adjudicated_label"].ne("") & ~adjudication["adjudicated_label"].isin(LABELS), "adjudicated_label"]))
    if invalid:
        raise ValueError(f"{split_name} adjudication has invalid labels: {invalid}")
    merged = required[["annotation_id", "comment_id"]].merge(
        adjudication[["annotation_id", "comment_id", "adjudicated_label", "adjudication_note"]],
        on=["annotation_id", "comment_id"],
        how="left",
        validate="one_to_one",
    )
    missing = merged["adjudicated_label"].astype(str).str.strip().eq("")
    if missing.any():
        examples = merged.loc[missing, "annotation_id"].head(10).tolist()
        raise ValueError(f"{split_name} adjudication incomplete for {int(missing.sum())} rows: {examples}")
    return merged[["annotation_id", "adjudicated_label", "adjudication_note"]]


def finalize_split(
    split_name: str,
    registry_path: Path,
    adjudication_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    registry = read_csv(registry_path)
    adjudication = read_first_sheet(adjudication_path)
    adjudicated = validate_adjudication(registry, adjudication, split_name)
    frame = registry.merge(adjudicated, on="annotation_id", how="left", suffixes=("", "_human"), validate="one_to_one")
    frame["adjudication_required"] = frame["adjudication_required"].map(parse_bool)
    frame["pre_adjudication_human_label"] = frame["pre_adjudication_human_label"].map(normalize_label)
    if "adjudicated_label_human" in frame.columns:
        frame["adjudicated_label"] = frame["adjudicated_label_human"]
    if "adjudication_note_human" in frame.columns:
        frame["adjudication_note"] = frame["adjudication_note_human"]
    frame["adjudicated_label"] = frame["adjudicated_label"].fillna("").map(normalize_label)
    frame["adjudication_note"] = frame["adjudication_note"].fillna("").astype(str)
    frame = frame.drop(columns=[column for column in ["adjudicated_label_human", "adjudication_note_human"] if column in frame.columns])
    frame["final_human_label"] = np.where(
        frame["adjudication_required"],
        frame["adjudicated_label"],
        frame["pre_adjudication_human_label"],
    )
    frame["final_label_source"] = np.where(
        frame["adjudication_required"],
        "human_adjudication",
        "annotator_agreement",
    )
    frame["annotation_status"] = np.where(
        frame["final_human_label"].isin(LABELS),
        "FINAL_HUMAN_LABEL",
        "INVALID_MISSING_FINAL_LABEL",
    )
    frame["evaluable_three_class"] = frame["final_human_label"].isin(THREE_CLASS)
    frame["exclusion_reason"] = np.select(
        [
            frame["final_human_label"].eq("Uncertain"),
            frame["final_human_label"].eq("No Text"),
            frame["final_human_label"].eq(""),
        ],
        ["Uncertain", "No Text", "Missing final human label"],
        default="",
    )
    if frame["annotation_status"].eq("INVALID_MISSING_FINAL_LABEL").any():
        raise ValueError(f"{split_name} has missing final labels after adjudication")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(output_path, index=False, encoding="utf-8-sig")
    distribution_stem = output_path.stem
    if distribution_stem.endswith("_registry"):
        distribution_stem = distribution_stem.removesuffix("_registry")
    distribution_path = output_path.with_name(f"{distribution_stem}_label_distribution.csv")
    distribution = (
        frame["final_human_label"]
        .value_counts()
        .reindex(LABELS, fill_value=0)
        .rename_axis("final_human_label")
        .reset_index(name="count")
    )
    distribution.to_csv(distribution_path, index=False, encoding="utf-8-sig")

    return {
        "rows": int(len(frame)),
        "unique_comment_ids": int(frame["comment_id"].nunique()),
        "disagreement_rows_adjudicated": int(frame["adjudication_required"].sum()),
        "evaluable_three_class_rows": int(frame["evaluable_three_class"].sum()),
        "label_counts": label_counts(frame),
        "output": output_path,
        "label_distribution": distribution_path,
        "dataset_hash": sha256_dataframe(frame, ["annotation_id", "comment_id", "final_human_label"]),
    }


def leakage_audit(dev: pd.DataFrame, locked: pd.DataFrame) -> dict[str, Any]:
    dev_ids = set(dev["comment_id"].astype(str))
    locked_ids = set(locked["comment_id"].astype(str))
    dev_text = set(dev["comment_text"].map(normalize_text)) - {""}
    locked_text = set(locked["comment_text"].map(normalize_text)) - {""}
    inj_count = int(
        dev["comment_id"].astype(str).str.contains("INJ", case=False, regex=False).sum()
        + locked["comment_id"].astype(str).str.contains("INJ", case=False, regex=False).sum()
    )
    return {
        "development_locked_comment_id_overlap": len(dev_ids & locked_ids),
        "development_locked_normalized_text_overlap": len(dev_text & locked_text),
        "inj_comment_id_count": inj_count,
        "development_duplicate_comment_ids": int(dev["comment_id"].duplicated().sum()),
        "locked_test_duplicate_comment_ids": int(locked["comment_id"].duplicated().sum()),
        "hard_leakage_pass": not (dev_ids & locked_ids) and not (dev_text & locked_text) and inj_count == 0,
    }


def main() -> None:
    selected_lock_adjudication = pick_adjudication_workbook(LOCK_ADJUDICATION, LOCK_ADJUDICATION_FALLBACK)
    synced_locked_workbook = sync_adjudication_workbook(selected_lock_adjudication, LOCK_ADJUDICATION)

    dev = finalize_split(
        "development",
        DEV_PRE_REGISTRY,
        DEV_ADJUDICATION,
        DEV_DIR / "sentiment_v5_development_final_registry.csv",
    )
    locked = finalize_split(
        "locked_test",
        LOCK_PRE_REGISTRY,
        LOCK_ADJUDICATION,
        LOCK_DIR / "sentiment_v5_locked_test_final_frozen.csv",
    )
    dev_frame = read_csv(Path(dev["output"]))
    locked_frame = read_csv(Path(locked["output"]))
    audit = leakage_audit(dev_frame, locked_frame)
    validation_rows = [
        {"check": "development_rows", "value": dev["rows"], "status": "PASS" if dev["rows"] == 1000 else "WARN"},
        {"check": "locked_test_rows", "value": locked["rows"], "status": "PASS" if locked["rows"] == 700 else "WARN"},
        {"check": "development_evaluable_three_class_rows", "value": dev["evaluable_three_class_rows"], "status": "PASS"},
        {"check": "locked_test_evaluable_three_class_rows", "value": locked["evaluable_three_class_rows"], "status": "PASS" if locked["evaluable_three_class_rows"] >= 600 else "WARN"},
        {"check": "comment_id_leakage", "value": audit["development_locked_comment_id_overlap"], "status": "PASS" if audit["development_locked_comment_id_overlap"] == 0 else "FAIL"},
        {"check": "normalized_text_leakage", "value": audit["development_locked_normalized_text_overlap"], "status": "PASS" if audit["development_locked_normalized_text_overlap"] == 0 else "FAIL"},
        {"check": "inj_comment_id_count", "value": audit["inj_comment_id_count"], "status": "PASS" if audit["inj_comment_id_count"] == 0 else "FAIL"},
    ]
    validation_path = DEV_DIR / "sentiment_v5_final_validation_report.csv"
    pd.DataFrame(validation_rows).to_csv(validation_path, index=False, encoding="utf-8-sig")
    leakage_path = LOCK_DIR / "sentiment_v5_final_split_leakage_audit.csv"
    pd.DataFrame(
        [
            {
                "check": "development_locked_comment_id_overlap",
                "value": audit["development_locked_comment_id_overlap"],
                "status": "PASS" if audit["development_locked_comment_id_overlap"] == 0 else "FAIL",
            },
            {
                "check": "development_locked_normalized_text_overlap",
                "value": audit["development_locked_normalized_text_overlap"],
                "status": "PASS" if audit["development_locked_normalized_text_overlap"] == 0 else "FAIL",
            },
            {
                "check": "inj_comment_id_count",
                "value": audit["inj_comment_id_count"],
                "status": "PASS" if audit["inj_comment_id_count"] == 0 else "FAIL",
            },
        ]
    ).to_csv(leakage_path, index=False, encoding="utf-8-sig")

    status = "SENTIMENT_V5_FINAL_HUMAN_LABELS_IMPORTED"
    if not audit["hard_leakage_pass"]:
        status = "SENTIMENT_V5_FINAL_IMPORT_BLOCKED_BY_LEAKAGE"
    manifest = {
        "status": status,
        "created_at_utc": utc_now(),
        "primary_label": "sentiment_toward_target",
        "development": dev,
        "locked_test_v5": locked,
        "leakage_audit": audit,
        "validation_report": validation_path,
        "split_leakage_audit": leakage_path,
        "adjudication_workbooks": {
            "development": DEV_ADJUDICATION,
            "locked_test_canonical": LOCK_ADJUDICATION,
            "locked_test_source_used": selected_lock_adjudication,
            "locked_test_canonical_synced_from_source": synced_locked_workbook,
        },
        "methodology": {
            "model_predictions_used_for_labels": False,
            "auto_final_labels_for_disagreements": False,
            "final_labels_from_human_agreement_or_adjudication_only": True,
            "locked_test_v5_used_for_training_or_selection": False,
            "v5_model_final": False,
            "locked_test_v5_sealed_until_preregistered_final_evaluation": True,
        },
    }
    final_manifest_path = DEV_DIR / "SENTIMENT_V5_FINAL_IMPORT_MANIFEST.json"
    write_json(final_manifest_path, manifest)

    import_manifest_path = DEV_DIR / "SENTIMENT_V5_IMPORT_MANIFEST.json"
    import_manifest = read_json(import_manifest_path) if import_manifest_path.exists() else {}
    import_manifest.update(
        {
            "status": "SENTIMENT_V5_FINAL_HUMAN_LABELS_FROZEN",
            "final_import_manifest": rel(final_manifest_path),
            "development_final_registry": rel(Path(dev["output"])),
            "locked_test_final_frozen": rel(Path(locked["output"])),
            "locked_test_v5_used_for_training_or_selection": False,
        }
    )
    write_json(import_manifest_path, import_manifest)

    lock_manifest_path = LOCK_DIR / "LOCKED_TEST_V5_FREEZE_MANIFEST.json"
    lock_manifest = read_json(lock_manifest_path)
    lock_manifest.update(
        {
            "status": "LOCKED_TEST_V5_FINAL_LABELS_FROZEN_SEALED",
            "final_label_freeze_status": "FROZEN_NOT_OPEN_FOR_MODEL_SELECTION",
            "final_label_registry": rel(Path(locked["output"])),
            "final_label_dataset_hash": locked["dataset_hash"],
            "final_label_counts": locked["label_counts"],
            "final_label_distribution": locked["label_counts"],
            "final_evaluable_three_class_rows": locked["evaluable_three_class_rows"],
            "final_label_freeze_timestamp_utc": utc_now(),
            "final_split_leakage_audit": rel(leakage_path),
            "locked_test_v5_used_for_training_or_selection": False,
            "model_selection_allowed_from_locked_test_v5": False,
            "locked_test_v5_may_only_be_evaluated_after_v5_model_freeze": True,
        }
    )
    write_json(lock_manifest_path, lock_manifest)
    print(json.dumps(to_jsonable(manifest), indent=2), flush=True)


if __name__ == "__main__":
    main()
