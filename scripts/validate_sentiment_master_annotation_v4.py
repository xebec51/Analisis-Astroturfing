from __future__ import annotations

import json
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.metrics import cohen_kappa_score


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output/rm2_sentiment/validation/human_master_v4"
MASTER = OUT_DIR / "human_annotation_master_v4.csv"
MANIFEST = OUT_DIR / "MASTER_ANNOTATION_V4_MANIFEST.json"
READINESS = OUT_DIR / "master_readiness_report.csv"
VALIDATION_REPORT = OUT_DIR / "master_validation_report.csv"
AGREEMENT_REPORT = OUT_DIR / "master_agreement_report.csv"

DEV_A1 = OUT_DIR / "sentiment_v4_development_annotator_1.xlsx"
DEV_A2 = OUT_DIR / "sentiment_v4_development_annotator_2.xlsx"
DEV_ADJ = OUT_DIR / "sentiment_v4_development_adjudication.xlsx"
LOCK_A1 = OUT_DIR / "sentiment_v4_locked_test_annotator_1.xlsx"
LOCK_A2 = OUT_DIR / "sentiment_v4_locked_test_annotator_2.xlsx"
LOCK_ADJ = OUT_DIR / "sentiment_v4_locked_test_adjudication.xlsx"

LABELS = ["Negative", "Neutral", "Positive", "Uncertain", "No Text"]
EVALUABLE = ["Negative", "Neutral", "Positive"]
PENDING_ROLES = ["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]
DEV_ROLES = ["HISTORICAL_DEVELOPMENT_FINAL", "DEVELOPMENT_NEW_PENDING", "DEVELOPMENT_NEW_FINAL"]
LOCKED_ROLES = ["LOCKED_TEST_NEW_PENDING", "LOCKED_TEST_NEW_FINAL"]
ROLE_VALUES = [
    "HISTORICAL_DEVELOPMENT_FINAL",
    "DEVELOPMENT_NEW_PENDING",
    "DEVELOPMENT_NEW_FINAL",
    "LEGACY_TEST_PROVENANCE",
    "LOCKED_TEST_NEW_PENDING",
    "LOCKED_TEST_NEW_FINAL",
    "EXCLUDED",
]
MASTER_COLUMNS = [
    "master_annotation_id",
    "comment_id",
    "video_id",
    "comment_text",
    "model_text",
    "product_category",
    "brand_or_video_context",
    "source_file",
    "source_version",
    "source_role",
    "annotation_role",
    "split_lock",
    "historical_final_label",
    "annotator_1_label",
    "annotator_2_label",
    "adjudicated_label",
    "final_human_label",
    "annotation_status",
    "disagreement_flag",
    "adjudication_required",
    "evaluable_three_class",
    "exclusion_reason",
    "text_cluster_id",
    "exact_duplicate_group_id",
    "near_duplicate_cluster_id",
    "sampling_stratum",
    "sampling_batch",
    "created_at",
    "notes",
]
ANNOTATOR_COLUMNS = [
    "annotation_id",
    "comment_id",
    "video_id",
    "product_category",
    "brand_or_video_context",
    "comment_text",
    "human_label",
    "confidence_annotation",
    "needs_context",
    "annotator_notes",
]
ADJUDICATION_COLUMNS = [
    "annotation_id",
    "comment_id",
    "video_id",
    "product_category",
    "brand_or_video_context",
    "comment_text",
    "annotator_1_label",
    "annotator_1_notes",
    "annotator_2_label",
    "annotator_2_notes",
    "adjudicated_label",
    "adjudication_notes",
]
BANNED = ["prediction", "probability", "threshold", "hcc", "actor", "goal", "model", "disagreement"]


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)


def is_inj(value: object) -> bool:
    text = "" if pd.isna(value) else str(value).strip().upper()
    return text.startswith("INJ") or "SYNTHETIC" in text


def gate(name: str, expected: object, observed: object, passed: bool, notes: str = "") -> dict[str, object]:
    return {"check": name, "expected": expected, "observed": observed, "passed": bool(passed), "notes": notes}


def hard_text_leakage(master: pd.DataFrame) -> int:
    failures = 0
    for _, group in master.groupby("text_cluster_id"):
        roles = set(group["annotation_role"])
        if roles & set(DEV_ROLES) and roles & set(LOCKED_ROLES):
            failures += 1
    return failures


def workbook_columns(path: Path) -> dict[str, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        out[ws.title] = [str(value) for value in row if value is not None]
    return out


def read_excel_sheet_with_columns(path: Path, required_columns: list[str]) -> pd.DataFrame:
    sheets = pd.read_excel(path, sheet_name=None, dtype=str, keep_default_na=False)
    required = set(required_columns)
    for frame in sheets.values():
        if required.issubset(set(frame.columns)):
            return frame
    raise AssertionError(f"{path.name} does not contain a sheet with columns: {required_columns}")


def annotator_workbook_ok(path: Path) -> tuple[bool, str]:
    columns_by_sheet = workbook_columns(path)
    data_sheet_count = 0
    for sheet, cols in columns_by_sheet.items():
        if sheet.upper() == "README":
            continue
        lower = [col.lower() for col in cols]
        if not set(ANNOTATOR_COLUMNS).issubset(cols):
            return False, f"{path.name}:{sheet} missing required annotator columns"
        data_sheet_count += 1
        banned_hits = [col for col in lower for token in BANNED if token in col and col not in {"annotation_id", "confidence_annotation", "annotator_notes"}]
        if banned_hits:
            return False, f"{path.name}:{sheet} contains banned columns {banned_hits}"
    if data_sheet_count != 1:
        return False, f"{path.name} should contain exactly one annotator data sheet"
    return True, ""


def read_annotator_labels(path: Path) -> pd.DataFrame:
    return read_excel_sheet_with_columns(path, ANNOTATOR_COLUMNS)


def agreement(a: pd.DataFrame, b: pd.DataFrame, scope: str) -> dict[str, object]:
    merged = a[["annotation_id", "human_label"]].merge(
        b[["annotation_id", "human_label"]],
        on="annotation_id",
        suffixes=("_1", "_2"),
        how="inner",
    )
    filled = merged.loc[merged["human_label_1"].isin(LABELS) & merged["human_label_2"].isin(LABELS)].copy()
    if filled.empty:
        return {
            "scope": scope,
            "n_compared": 0,
            "raw_agreement": "",
            "cohen_kappa": "",
            "status": "PENDING_HUMAN_LABELS",
        }
    raw = float(filled["human_label_1"].eq(filled["human_label_2"]).mean())
    kappa = float(cohen_kappa_score(filled["human_label_1"], filled["human_label_2"], labels=LABELS))
    return {
        "scope": scope,
        "n_compared": int(len(filled)),
        "raw_agreement": raw,
        "cohen_kappa": kappa,
        "status": "PASS" if raw >= 0.85 and kappa >= 0.80 else "REVIEW_LOW_AGREEMENT",
    }


def existing_adjudication_labels(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["annotation_id", "adjudicated_label", "adjudication_notes"])
    try:
        frame = read_excel_sheet_with_columns(path, ["annotation_id"])
    except AssertionError:
        return pd.DataFrame(columns=["annotation_id", "adjudicated_label", "adjudication_notes"])
    for col in ["adjudicated_label", "adjudication_notes"]:
        if col not in frame.columns:
            frame[col] = ""
    return frame[["annotation_id", "adjudicated_label", "adjudication_notes"]]


def disagreement_rows(master: pd.DataFrame, a1: pd.DataFrame, a2: pd.DataFrame, role: str, existing_adj: pd.DataFrame) -> pd.DataFrame:
    base = master.loc[
        master["annotation_role"].eq(role),
        ["master_annotation_id", "comment_id", "video_id", "product_category", "brand_or_video_context", "comment_text"],
    ].rename(columns={"master_annotation_id": "annotation_id"})
    left = a1[["annotation_id", "human_label", "annotator_notes"]].rename(
        columns={"human_label": "annotator_1_label", "annotator_notes": "annotator_1_notes"}
    )
    right = a2[["annotation_id", "human_label", "annotator_notes"]].rename(
        columns={"human_label": "annotator_2_label", "annotator_notes": "annotator_2_notes"}
    )
    merged = base.merge(left, on="annotation_id", how="left").merge(right, on="annotation_id", how="left")
    for col in ["annotator_1_label", "annotator_2_label", "annotator_1_notes", "annotator_2_notes"]:
        merged[col] = merged[col].fillna("")
    filled = merged["annotator_1_label"].isin(LABELS) & merged["annotator_2_label"].isin(LABELS)
    disagreements = merged.loc[filled & merged["annotator_1_label"].ne(merged["annotator_2_label"])].copy()
    if disagreements.empty:
        return pd.DataFrame(columns=ADJUDICATION_COLUMNS)
    keep_adj = existing_adj.loc[existing_adj["annotation_id"].isin(disagreements["annotation_id"])].copy()
    disagreements = disagreements.merge(keep_adj, on="annotation_id", how="left")
    for col in ["adjudicated_label", "adjudication_notes"]:
        disagreements[col] = disagreements[col].fillna("")
    return disagreements[ADJUDICATION_COLUMNS]


def adjudication_workbook_matches(path: Path, sheet_name: str, disagreements: pd.DataFrame) -> bool:
    if not path.exists():
        return False
    if sheet_name not in load_workbook(path, read_only=True).sheetnames:
        return False
    try:
        existing = read_excel_sheet_with_columns(path, ADJUDICATION_COLUMNS)
    except AssertionError:
        return False
    existing = existing.reindex(columns=ADJUDICATION_COLUMNS).fillna("").astype(str)
    expected = disagreements.reindex(columns=ADJUDICATION_COLUMNS).fillna("").astype(str).reset_index(drop=True)
    existing = existing.reset_index(drop=True)
    if len(existing) != len(expected):
        return False
    return existing.equals(expected)


def write_adjudication_workbook(path: Path, sheet_name: str, disagreements: pd.DataFrame) -> None:
    if adjudication_workbook_matches(path, sheet_name, disagreements):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["Adjudication workbook generated from annotator disagreements only."])
    ws.append(["Do not use model predictions, probabilities, HCC, actor type, or goal orientation as labels."])
    data = wb.create_sheet(sheet_name)
    data.append(ADJUDICATION_COLUMNS)
    for row in disagreements.itertuples(index=False):
        data.append(list(row))
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions
    for cell in data[1]:
        cell.font = Font(bold=True)
    widths = {
        "A": 18,
        "B": 24,
        "C": 22,
        "D": 20,
        "E": 28,
        "F": 70,
        "G": 18,
        "H": 30,
        "I": 18,
        "J": 30,
        "K": 20,
        "L": 36,
    }
    for col, width in widths.items():
        data.column_dimensions[col].width = width
    for row in data.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if data.max_row >= 2:
        label_validation = DataValidation(type="list", formula1='"Negative,Neutral,Positive,Uncertain,No Text"', allow_blank=True)
        data.add_data_validation(label_validation)
        label_validation.add(f"K2:K{data.max_row}")
    wb.save(path)


def main() -> None:
    master = read_csv(MASTER)
    checks = []
    checks.append(gate("schema_columns", MASTER_COLUMNS, list(master.columns), list(master.columns) == MASTER_COLUMNS))
    checks.append(gate("comment_id_unique", 0, int(master["comment_id"].duplicated().sum()), int(master["comment_id"].duplicated().sum()) == 0))
    checks.append(gate("role_vocabulary", ROLE_VALUES, sorted(master["annotation_role"].unique()), set(master["annotation_role"]).issubset(set(ROLE_VALUES))))

    pending = master.loc[master["annotation_role"].isin(PENDING_ROLES)].copy()
    non_pending = master.loc[~master["annotation_role"].isin(PENDING_ROLES)].copy()
    checks.append(gate("pending_total", 2000, len(pending), len(pending) == 2000))
    checks.append(gate("development_pending", 1300, int(master["annotation_role"].eq("DEVELOPMENT_NEW_PENDING").sum()), int(master["annotation_role"].eq("DEVELOPMENT_NEW_PENDING").sum()) == 1300))
    checks.append(gate("locked_test_pending", 700, int(master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING").sum()), int(master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING").sum()) == 700))
    checks.append(gate("pending_inj_count", 0, int(pending["comment_id"].map(is_inj).sum()), int(pending["comment_id"].map(is_inj).sum()) == 0))
    checks.append(gate("pending_final_labels_blank", True, bool(pending["final_human_label"].eq("").all()), bool(pending["final_human_label"].eq("").all())))
    checks.append(gate("pending_adjudicated_blank", True, bool(pending["adjudicated_label"].eq("").all()), bool(pending["adjudicated_label"].eq("").all())))
    label_values = set(non_pending["final_human_label"]) - {""}
    checks.append(gate("final_label_vocabulary", LABELS, sorted(label_values), label_values.issubset(set(LABELS))))

    dev_ids = set(master.loc[master["annotation_role"].isin(DEV_ROLES), "comment_id"])
    lock_ids = set(master.loc[master["annotation_role"].isin(LOCKED_ROLES), "comment_id"])
    checks.append(gate("hard_comment_id_leakage", 0, len(dev_ids & lock_ids), len(dev_ids & lock_ids) == 0))
    text_leakage = hard_text_leakage(master)
    checks.append(gate("hard_text_cluster_leakage", 0, text_leakage, text_leakage == 0))

    legacy_dev = master.loc[
        master["annotation_role"].isin(DEV_ROLES)
        & master["source_file"].str.contains("locked_test|final_test", case=False, regex=True, na=False)
    ]
    checks.append(gate("legacy_test_not_in_development", 0, len(legacy_dev), len(legacy_dev) == 0))
    locked_in_dev = master.loc[master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING") & master["split_lock"].str.contains("DEVELOPMENT", case=False, na=False)]
    checks.append(gate("new_locked_test_not_in_development", 0, len(locked_in_dev), len(locked_in_dev) == 0))

    workbook_results = []
    for path in [DEV_A1, DEV_A2, LOCK_A1, LOCK_A2]:
        ok, notes = annotator_workbook_ok(path)
        workbook_results.append((path.name, ok, notes))
        checks.append(gate(f"annotator_workbook_no_predictions_{path.name}", True, ok, ok, notes))

    dev_a1 = read_annotator_labels(DEV_A1)
    dev_a2 = read_annotator_labels(DEV_A2)
    lock_a1 = read_annotator_labels(LOCK_A1)
    lock_a2 = read_annotator_labels(LOCK_A2)

    agreements = pd.DataFrame(
        [
            agreement(dev_a1, dev_a2, "development_pending"),
            agreement(lock_a1, lock_a2, "locked_test_pending"),
        ]
    )
    agreements.to_csv(AGREEMENT_REPORT, index=False, encoding="utf-8-sig")
    dev_disagreements = disagreement_rows(master, dev_a1, dev_a2, "DEVELOPMENT_NEW_PENDING", existing_adjudication_labels(DEV_ADJ))
    lock_disagreements = disagreement_rows(master, lock_a1, lock_a2, "LOCKED_TEST_NEW_PENDING", existing_adjudication_labels(LOCK_ADJ))
    write_adjudication_workbook(DEV_ADJ, "DEVELOPMENT_ADJUDICATION", dev_disagreements)
    write_adjudication_workbook(LOCK_ADJ, "LOCKED_TEST_ADJUDICATION", lock_disagreements)

    report = pd.DataFrame(checks)
    final_ready = bool(report["passed"].all())
    report.to_csv(VALIDATION_REPORT, index=False, encoding="utf-8-sig")

    readiness = read_csv(READINESS)
    readiness = readiness.loc[~readiness["metric"].isin(["validation_status", "agreement_status"])].copy()
    readiness = pd.concat(
        [
            readiness,
            pd.DataFrame(
                [
                    {
                        "metric": "validation_status",
                        "value": "MASTER_ANNOTATION_V4_READY_FOR_HUMAN_LABELING" if final_ready else "MASTER_ANNOTATION_V4_VALIDATION_FAILED",
                        "passed": final_ready,
                        "notes": "Validation script result.",
                    },
                    {
                        "metric": "agreement_status",
                        "value": ";".join(agreements["status"].astype(str)),
                        "passed": True,
                        "notes": "Agreement is pending until annotator workbooks are completed.",
                    },
                ]
            ),
        ],
        ignore_index=True,
        sort=False,
    )
    readiness.to_csv(READINESS, index=False, encoding="utf-8-sig")

    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    manifest["validation_status"] = "MASTER_ANNOTATION_V4_READY_FOR_HUMAN_LABELING" if final_ready else "MASTER_ANNOTATION_V4_VALIDATION_FAILED"
    manifest["validation_report"] = "output/rm2_sentiment/validation/human_master_v4/master_validation_report.csv"
    manifest["agreement_report"] = "output/rm2_sentiment/validation/human_master_v4/master_agreement_report.csv"
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if not final_ready:
        failed = report.loc[~report["passed"]]
        raise AssertionError(f"Master V4 validation failed:\n{failed.to_string(index=False)}")
    print("MASTER_ANNOTATION_V4_READY_FOR_HUMAN_LABELING")
    print(f"checks={len(report)} agreement={agreements['status'].tolist()}")


if __name__ == "__main__":
    main()
