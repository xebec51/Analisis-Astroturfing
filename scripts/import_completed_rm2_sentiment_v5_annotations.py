from __future__ import annotations

import hashlib
import json
import math
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.metrics import cohen_kappa_score


ROOT = Path(__file__).resolve().parents[1]
DEV_DIR = ROOT / "output/rm2_sentiment/validation/human_v5"
LOCK_DIR = ROOT / "output/rm2_sentiment/validation/human_v5_locked_test"

LABELS = ["Negative", "Neutral", "Positive", "Uncertain", "No Text"]
THREE_CLASS = ["Negative", "Neutral", "Positive"]
FLAGS = ["Yes", "No"]
TARGET_BRANDS = ["Azarine", "Daviena", "Maryame", "The Originote", "Other", "Not clear"]
CONFIDENCE = ["High", "Medium", "Low"]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def rel(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return str(path)


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(to_jsonable(payload), indent=2), encoding="utf-8")


def to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): to_jsonable(v) for k, v in value.items()}
    if isinstance(value, list | tuple):
        return [to_jsonable(v) for v in value]
    if isinstance(value, Path):
        return rel(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_dataframe(frame: pd.DataFrame, columns: list[str]) -> str:
    data = frame[columns].sort_values(columns).to_csv(index=False).encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def read_first_sheet(path: Path) -> pd.DataFrame:
    sheet = pd.ExcelFile(path).sheet_names[0]
    frame = pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
    frame.columns = [str(column).strip() for column in frame.columns]
    return frame


def normalize_label(value: object) -> str:
    text = "" if value is None or pd.isna(value) else str(value).strip()
    aliases = {
        "negative": "Negative",
        "neutral": "Neutral",
        "positive": "Positive",
        "uncertain": "Uncertain",
        "no text": "No Text",
        "notext": "No Text",
        "no_text": "No Text",
    }
    return aliases.get(text.lower(), text)


def validate_values(frame: pd.DataFrame, columns: dict[str, list[str]], source_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for column, allowed in columns.items():
        if column not in frame.columns:
            rows.append({"source": source_name, "column": column, "invalid_count": -1, "invalid_values": "MISSING_COLUMN"})
            continue
        values = frame[column].astype(str).str.strip()
        invalid = sorted(set(values.loc[values.ne("") & ~values.isin(allowed)]))
        rows.append(
            {
                "source": source_name,
                "column": column,
                "invalid_count": len(invalid),
                "invalid_values": "; ".join(invalid),
            }
        )
    return rows


def agreement_stats(left: pd.Series, right: pd.Series) -> dict[str, Any]:
    mask = left.astype(str).str.strip().ne("") & right.astype(str).str.strip().ne("")
    lvals = left.loc[mask].astype(str)
    rvals = right.loc[mask].astype(str)
    total = int(mask.sum())
    agreements = int(lvals.eq(rvals).sum())
    return {
        "compared_rows": total,
        "agreements": agreements,
        "disagreements": total - agreements,
        "raw_agreement": agreements / total if total else np.nan,
        "cohen_kappa": float(cohen_kappa_score(lvals, rvals, labels=LABELS)) if total else np.nan,
    }


def merge_annotators(a1: pd.DataFrame, a2: pd.DataFrame, role: str) -> pd.DataFrame:
    for frame, annotator in [(a1, "annotator_1"), (a2, "annotator_2")]:
        for label_col in ["sentiment_overall", "sentiment_toward_target"]:
            frame[label_col] = frame[label_col].map(normalize_label)
        frame[f"{annotator}_source_row"] = np.arange(len(frame)) + 2
    required = ["annotation_id", "comment_id", "sentiment_toward_target", "sentiment_overall"]
    for required_col in required:
        if required_col not in a1.columns or required_col not in a2.columns:
            raise ValueError(f"Missing required column: {required_col}")
    merged = a1.merge(
        a2,
        on=["annotation_id", "comment_id"],
        how="outer",
        suffixes=("_annotator_1", "_annotator_2"),
        indicator=True,
        validate="one_to_one",
    )
    if not merged["_merge"].eq("both").all():
        raise ValueError(f"{role} annotator files have mismatched annotation/comment IDs.")
    base_cols = [
        "annotation_id",
        "comment_id",
        "video_id_annotator_1",
        "product_category_annotator_1",
        "brand_or_video_context_annotator_1",
        "comment_text_annotator_1",
        "parent_comment_text_annotator_1",
    ]
    out = merged[base_cols].rename(
        columns={
            "video_id_annotator_1": "video_id",
            "product_category_annotator_1": "product_category",
            "brand_or_video_context_annotator_1": "brand_or_video_context",
            "comment_text_annotator_1": "comment_text",
            "parent_comment_text_annotator_1": "parent_comment_text",
        }
    )
    out["annotation_role"] = role
    out["annotator_1_sentiment_overall"] = merged["sentiment_overall_annotator_1"]
    out["annotator_2_sentiment_overall"] = merged["sentiment_overall_annotator_2"]
    out["annotator_1_sentiment_toward_target"] = merged["sentiment_toward_target_annotator_1"]
    out["annotator_2_sentiment_toward_target"] = merged["sentiment_toward_target_annotator_2"]
    out["annotator_1_target_brand"] = merged.get("target_brand_annotator_1", "")
    out["annotator_2_target_brand"] = merged.get("target_brand_annotator_2", "")
    out["annotator_1_notes"] = merged.get("annotator_notes_annotator_1", "")
    out["annotator_2_notes"] = merged.get("annotator_notes_annotator_2", "")
    out["sentiment_toward_target_agreement"] = out["annotator_1_sentiment_toward_target"].eq(out["annotator_2_sentiment_toward_target"])
    out["sentiment_overall_agreement"] = out["annotator_1_sentiment_overall"].eq(out["annotator_2_sentiment_overall"])
    out["adjudication_required"] = ~out["sentiment_toward_target_agreement"]
    out["adjudicated_label"] = ""
    out["adjudication_note"] = ""
    out["pre_adjudication_human_label"] = np.where(
        out["sentiment_toward_target_agreement"],
        out["annotator_1_sentiment_toward_target"],
        "",
    )
    out["evaluable_three_class_pre_adjudication"] = out["pre_adjudication_human_label"].isin(THREE_CLASS)
    out["annotation_status"] = np.where(
        out["adjudication_required"],
        "PENDING_HUMAN_ADJUDICATION",
        "AGREEMENT_READY",
    )
    return out


def write_adjudication_workbook(path: Path, disagreements: pd.DataFrame, sheet_name: str) -> None:
    columns = [
        "annotation_id",
        "comment_id",
        "video_id",
        "product_category",
        "brand_or_video_context",
        "comment_text",
        "parent_comment_text",
        "annotator_1_sentiment_toward_target",
        "annotator_2_sentiment_toward_target",
        "annotator_1_sentiment_overall",
        "annotator_2_sentiment_overall",
        "annotator_1_notes",
        "annotator_2_notes",
        "adjudicated_label",
        "adjudication_note",
    ]
    frame = disagreements.copy()
    for column in columns:
        if column not in frame.columns:
            frame[column] = ""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame[columns].to_excel(writer, index=False, sheet_name=sheet_name)
        pd.DataFrame(
            {
                "instruction": [
                    "Isi hanya adjudicated_label dan adjudication_note.",
                    "Pilih adjudicated_label dari Negative, Neutral, Positive, Uncertain, No Text.",
                    "Jangan memakai prediksi model, HCC, actor type, atau goal orientation.",
                    "Label utama V5 adalah sentiment_toward_target.",
                ]
            }
        ).to_excel(writer, index=False, sheet_name="README")
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column, width in {"A": 18, "B": 24, "F": 72, "G": 48, "N": 20, "O": 48}.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
    max_row = max(ws.max_row, 2)
    dv = DataValidation(type="list", formula1=f'"{",".join(LABELS)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"N2:N{max_row}")
    wb.save(path)


def process_split(name: str, directory: Path, role: str, a1_name: str, a2_name: str, adj_name: str) -> dict[str, Any]:
    a1_path = directory / a1_name
    a2_path = directory / a2_name
    adj_path = directory / adj_name
    a1 = read_first_sheet(a1_path)
    a2 = read_first_sheet(a2_path)
    validation_rows = []
    allowed_columns = {
        "sentiment_overall": LABELS,
        "sentiment_toward_target": LABELS,
        "target_brand": TARGET_BRANDS,
        "mixed_sentiment_flag": FLAGS,
        "question_flag": FLAGS,
        "comparison_brand_flag": FLAGS,
        "sarcasm_or_irony_flag": FLAGS,
        "insufficient_context_flag": FLAGS,
        "confidence_annotation": CONFIDENCE,
    }
    validation_rows.extend(validate_values(a1, allowed_columns, f"{name}_annotator_1"))
    validation_rows.extend(validate_values(a2, allowed_columns, f"{name}_annotator_2"))
    pd.DataFrame(validation_rows).to_csv(directory / f"sentiment_v5_{name}_value_validation.csv", index=False, encoding="utf-8-sig")
    registry = merge_annotators(a1, a2, role)
    disagreements = registry.loc[registry["adjudication_required"]].copy()
    registry.to_csv(directory / f"sentiment_v5_{name}_pre_adjudication_registry.csv", index=False, encoding="utf-8-sig")
    disagreements.to_csv(directory / f"sentiment_v5_{name}_disagreements.csv", index=False, encoding="utf-8-sig")
    write_adjudication_workbook(adj_path, disagreements, f"{name.upper()}_ADJUDICATION")
    target_stats = agreement_stats(registry["annotator_1_sentiment_toward_target"], registry["annotator_2_sentiment_toward_target"])
    overall_stats = agreement_stats(registry["annotator_1_sentiment_overall"], registry["annotator_2_sentiment_overall"])
    report = pd.DataFrame(
        [
            {"field": "sentiment_toward_target", **target_stats},
            {"field": "sentiment_overall", **overall_stats},
        ]
    )
    report.to_csv(directory / f"sentiment_v5_{name}_agreement_report.csv", index=False, encoding="utf-8-sig")
    return {
        "name": name,
        "rows": int(len(registry)),
        "agreement_rows": int((~registry["adjudication_required"]).sum()),
        "disagreement_rows": int(registry["adjudication_required"].sum()),
        "pre_adjudication_three_class_rows": int(registry["evaluable_three_class_pre_adjudication"].sum()),
        "sentiment_toward_target": target_stats,
        "sentiment_overall": overall_stats,
        "registry": directory / f"sentiment_v5_{name}_pre_adjudication_registry.csv",
        "disagreements": directory / f"sentiment_v5_{name}_disagreements.csv",
        "agreement_report": directory / f"sentiment_v5_{name}_agreement_report.csv",
        "adjudication_workbook": adj_path,
        "annotator_hashes": {
            a1_name: sha256_file(a1_path),
            a2_name: sha256_file(a2_path),
        },
        "pre_adjudication_hash": sha256_dataframe(registry, ["annotation_id", "comment_id", "annotator_1_sentiment_toward_target", "annotator_2_sentiment_toward_target"]),
    }


def main() -> None:
    dev = process_split(
        "development",
        DEV_DIR,
        "DEVELOPMENT_V5_NEW_PENDING",
        "sentiment_v5_development_annotator_1.xlsx",
        "sentiment_v5_development_annotator_2.xlsx",
        "sentiment_v5_development_adjudication.xlsx",
    )
    locked = process_split(
        "locked_test",
        LOCK_DIR,
        "LOCKED_TEST_V5_NEW_PENDING",
        "sentiment_v5_locked_test_annotator_1.xlsx",
        "sentiment_v5_locked_test_annotator_2.xlsx",
        "sentiment_v5_locked_test_adjudication.xlsx",
    )
    status = "SENTIMENT_V5_PENDING_HUMAN_ADJUDICATION" if dev["disagreement_rows"] or locked["disagreement_rows"] else "SENTIMENT_V5_READY_FOR_FINAL_FREEZE"
    manifest = {
        "status": status,
        "created_at_utc": utc_now(),
        "primary_label": "sentiment_toward_target",
        "development": dev,
        "locked_test": locked,
        "methodology": {
            "auto_final_labels_for_disagreements": False,
            "model_predictions_used_for_labels": False,
            "locked_test_v5_used_for_training_or_selection": False,
            "adjudication_required_before_v5_training_or_locked_evaluation": bool(dev["disagreement_rows"] or locked["disagreement_rows"]),
            "do_not_train_until_development_adjudication_complete": True,
            "do_not_open_locked_test_for_model_selection": True,
        },
    }
    write_json(DEV_DIR / "SENTIMENT_V5_IMPORT_MANIFEST.json", manifest)
    lock_manifest_path = LOCK_DIR / "LOCKED_TEST_V5_FREEZE_MANIFEST.json"
    lock_manifest = read_json(lock_manifest_path)
    lock_manifest.update(
        {
            "status": "LOCKED_TEST_V5_PENDING_HUMAN_ADJUDICATION",
            "final_label_freeze_status": "BLOCKED_PENDING_ADJUDICATION",
            "annotator_label_import_status": "ANNOTATOR_1_AND_2_IMPORTED",
            "pre_adjudication_registry": rel(locked["registry"]),
            "disagreement_rows": locked["disagreement_rows"],
            "agreement_rows": locked["agreement_rows"],
            "final_label_dataset_hash": "",
            "locked_test_v5_used_for_training_or_selection": False,
        }
    )
    write_json(lock_manifest_path, lock_manifest)
    print(json.dumps(to_jsonable({"status": status, "development": dev, "locked_test": locked}), indent=2), flush=True)


if __name__ == "__main__":
    main()
