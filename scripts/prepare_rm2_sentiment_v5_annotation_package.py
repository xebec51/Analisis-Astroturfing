from __future__ import annotations

import hashlib
import json
import math
import re
import unicodedata
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DATASET = ROOT / "dataset.csv"
V2_FINAL = ROOT / "output/rm2_sentiment/final/comment_sentiment_v2_observational.csv"
V4_DEV = ROOT / "output/rm2_sentiment/validation/human_master_v4/sentiment_v4_development_final_registry.csv"
V4_LOCKED = ROOT / "output/rm2_sentiment/validation/human_master_v4/sentiment_v4_locked_test_final_frozen.csv"
HUMAN_V5_DIR = ROOT / "output/rm2_sentiment/validation/human_v5"
LOCKED_V5_DIR = ROOT / "output/rm2_sentiment/validation/human_v5_locked_test"
RANDOM_SEED = 20260728

LABEL_CHOICES = ["Negative", "Neutral", "Positive", "Uncertain", "No Text"]
FLAG_CHOICES = ["Yes", "No"]
CONFIDENCE_CHOICES = ["High", "Medium", "Low"]
TARGET_BRANDS = ["Azarine", "Daviena", "Maryame", "The Originote", "Other", "Not clear"]

ANNOTATION_COLUMNS = [
    "annotation_id",
    "comment_id",
    "video_id",
    "product_category",
    "brand_or_video_context",
    "comment_text",
    "parent_comment_text",
    "sentiment_overall",
    "sentiment_toward_target",
    "target_brand",
    "mixed_sentiment_flag",
    "question_flag",
    "comparison_brand_flag",
    "sarcasm_or_irony_flag",
    "insufficient_context_flag",
    "confidence_annotation",
    "annotator_notes",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def rel(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return str(path)


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(to_jsonable(payload), indent=2), encoding="utf-8")


def to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): to_jsonable(v) for k, v in value.items()}
    if isinstance(value, list | tuple):
        return [to_jsonable(v) for v in value]
    if isinstance(value, Path):
        return rel(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_dataframe(frame: pd.DataFrame, columns: list[str]) -> str:
    data = frame[columns].sort_values(columns).to_csv(index=False).encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def normalize_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null", "<na>", "[deleted]", "deleted"}:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"https?://\S+|www\.\S+", " HTTPURL ", text)
    text = re.sub(r"@\w+", " @USER ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def brand_from_category(value: object) -> str:
    text = str(value).lower()
    if "azarine" in text:
        return "Azarine"
    if "daviena" in text:
        return "Daviena"
    if "maryame" in text:
        return "Maryame"
    if "originote" in text:
        return "The Originote"
    return "Not clear"


def month_bucket(value: object) -> str:
    text = str(value)
    match = re.match(r"(\d{4}-\d{2})", text)
    return match.group(1) if match else "unknown"


def text_word_count(value: object) -> int:
    return len(normalize_text(value).split())


def question_flag(text: str) -> bool:
    value = normalize_text(text)
    return bool(re.search(r"\?|\\b(apa|apakah|gimana|bagaimana|bisa|boleh|aman|berapa|kenapa|cara|cocok gak|cocok nggak)\\b", value))


def comparison_brand_flag(text: str) -> bool:
    value = normalize_text(text)
    brands = ["azarine", "daviena", "maryame", "originote", "avoskin", "skintific", "somethinc", "wardah"]
    return sum(brand in value for brand in brands) >= 2 or bool(re.search(r"\\b(vs|versus|dibanding|mending|daripada)\\b", value))


def mixed_sentiment_flag(text: str) -> bool:
    value = normalize_text(text)
    positive = bool(re.search(r"\\b(bagus|cocok|worth|suka|puas|mantap|lembab|glowing|rekomen|aman|mulus)\\b", value))
    negative = bool(re.search(r"\\b(gak|nggak|tidak|kurang|jerawat|bruntus|perih|iritasi|mahal|bau|lengket|takut)\\b", value))
    return positive and negative


def emoji_or_slang_flag(text: str) -> bool:
    value = str(text)
    slang = bool(re.search(r"\\b(bgt|bangettt|wkwk|anjir|sih|dong|pls|plis|ga|gak|nggak|nih|deh)\\b", value.lower()))
    emoji = any(ord(ch) > 10000 for ch in value)
    return emoji or slang


def negative_word_positive_candidate(text: str) -> bool:
    value = normalize_text(text)
    return bool(re.search(r"\\b(gak|nggak|tidak|kurang|jerawat|perih|iritasi|mahal|bau|lengket)\\b", value)) and bool(
        re.search(r"\\b(cocok|bagus|aman|worth|suka|lembab|rekomen|puas)\\b", value)
    )


def assign_strata(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    probs = out[["probability_negative", "probability_neutral", "probability_positive"]].apply(pd.to_numeric, errors="coerce")
    out["v2_confidence"] = probs.max(axis=1)
    out["v2_margin"] = probs.apply(lambda row: float(np.sort(row.to_numpy(dtype=float))[-1] - np.sort(row.to_numpy(dtype=float))[-2]), axis=1)
    out["text_norm"] = out["text"].map(normalize_text)
    out["question_flag_sampling"] = out["text"].map(question_flag)
    out["comparison_brand_flag_sampling"] = out["text"].map(comparison_brand_flag)
    out["mixed_sentiment_flag_sampling"] = out["text"].map(mixed_sentiment_flag)
    out["emoji_or_slang_flag_sampling"] = out["text"].map(emoji_or_slang_flag)
    out["short_comment_flag_sampling"] = out["text"].map(lambda value: 0 < text_word_count(value) <= 4)
    out["negative_word_positive_candidate_sampling"] = out["text"].map(negative_word_positive_candidate)
    out["v2_abstained_sampling"] = out["final_sentiment_label"].eq("Uncertain")
    out["low_confidence_sampling"] = out["v2_confidence"].lt(0.55) | out["v2_margin"].lt(0.15)
    return out


def sample_by_plan(frame: pd.DataFrame, plan: list[tuple[str, str, int]], prefix: str) -> pd.DataFrame:
    rng = np.random.default_rng(RANDOM_SEED + (17 if prefix.startswith("LOCK") else 0))
    selected_ids: set[str] = set()
    selected_text: set[str] = set()
    chunks: list[pd.DataFrame] = []
    pool = frame.sample(frac=1.0, random_state=RANDOM_SEED).reset_index(drop=True)
    for stratum, column, target in plan:
        if column == "random_control":
            candidates = pool.copy()
        else:
            candidates = pool.loc[pool[column].astype(bool)].copy()
        candidates = candidates.loc[
            ~candidates["comment_id"].isin(selected_ids)
            & ~candidates["text_norm"].isin(selected_text)
            & candidates["text_norm"].ne("")
        ]
        if len(candidates) > target:
            candidates = candidates.sample(n=target, random_state=int(rng.integers(1, 1_000_000)))
        candidates = candidates.copy()
        candidates["sampling_stratum"] = stratum
        selected_ids.update(candidates["comment_id"].astype(str))
        selected_text.update(candidates["text_norm"].astype(str))
        chunks.append(candidates)
    sampled = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=pool.columns)
    shortfall = sum(target for _, _, target in plan) - len(sampled)
    if shortfall > 0:
        filler = pool.loc[
            ~pool["comment_id"].isin(selected_ids)
            & ~pool["text_norm"].isin(selected_text)
            & pool["text_norm"].ne("")
        ].copy()
        filler = filler.sample(n=shortfall, random_state=int(rng.integers(1, 1_000_000))).copy()
        filler["sampling_stratum"] = "random_shortfall_fill"
        sampled = pd.concat([sampled, filler], ignore_index=True)
    sampled = sampled.sample(frac=1.0, random_state=int(rng.integers(1, 1_000_000))).reset_index(drop=True)
    sampled["annotation_id"] = [f"{prefix}{idx:04d}" for idx in range(1, len(sampled) + 1)]
    return sampled


def candidate_columns(sampled: pd.DataFrame, role: str) -> pd.DataFrame:
    parent_lookup = sampled.set_index("comment_id")["text"].to_dict()
    out = pd.DataFrame(
        {
            "annotation_id": sampled["annotation_id"],
            "comment_id": sampled["comment_id"],
            "video_id": sampled["video_id"],
            "product_category": sampled["product_category"],
            "brand_or_video_context": sampled["product_category"],
            "comment_text": sampled["text"],
            "parent_comment_id": sampled["parent_comment_id"],
            "parent_comment_text": sampled["parent_comment_id"].map(parent_lookup).fillna(""),
            "target_brand_prefill": sampled["product_category"].map(brand_from_category),
            "sampling_stratum": sampled["sampling_stratum"],
            "sampling_role": role,
            "sampling_batch": "sentiment_v5_20260728",
            "created_at_utc": utc_now(),
            "sentiment_overall": "",
            "sentiment_toward_target": "",
            "target_brand": "",
            "mixed_sentiment_flag": "",
            "question_flag": "",
            "comparison_brand_flag": "",
            "sarcasm_or_irony_flag": "",
            "insufficient_context_flag": "",
            "annotator_1": "",
            "annotator_2": "",
            "adjudicated_label": "",
            "adjudication_note": "",
            "final_human_label": "",
            "evaluable_three_class": "",
        }
    )
    return out


def annotator_view(candidates: pd.DataFrame) -> pd.DataFrame:
    view = candidates.rename(columns={"target_brand_prefill": "brand_or_video_context"}).copy()
    view["brand_or_video_context"] = candidates["target_brand_prefill"]
    for column in ANNOTATION_COLUMNS:
        if column not in view.columns:
            view[column] = ""
    return view[ANNOTATION_COLUMNS]


def write_workbook(path: Path, frame: pd.DataFrame, sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name=sheet_name)
        pd.DataFrame(
            {
                "field": [
                    "sentiment_overall",
                    "sentiment_toward_target",
                    "target_brand",
                    "mixed_sentiment_flag",
                    "question_flag",
                    "comparison_brand_flag",
                    "sarcasm_or_irony_flag",
                    "insufficient_context_flag",
                ],
                "instruction": [
                    "General sentiment orientation of the comment text.",
                    "Primary RM2 model label: sentiment toward target brand/product/video context.",
                    "Target brand being evaluated, if identifiable.",
                    "Yes if the comment contains meaningful positive and negative elements.",
                    "Yes if the comment is primarily a question.",
                    "Yes if it compares two or more brands/products.",
                    "Yes if sarcasm or irony materially affects interpretation.",
                    "Yes if context is insufficient for a confident target-aware label.",
                ],
            }
        ).to_excel(writer, index=False, sheet_name="README")
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = {
        "A": 18,
        "B": 24,
        "C": 22,
        "D": 20,
        "E": 18,
        "F": 70,
        "G": 45,
        "H": 18,
        "I": 22,
        "J": 20,
        "Q": 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
    validations = {
        "H": LABEL_CHOICES,
        "I": LABEL_CHOICES,
        "J": TARGET_BRANDS,
        "K": FLAG_CHOICES,
        "L": FLAG_CHOICES,
        "M": FLAG_CHOICES,
        "N": FLAG_CHOICES,
        "O": FLAG_CHOICES,
        "P": CONFIDENCE_CHOICES,
    }
    max_row = max(ws.max_row, 2)
    for col, choices in validations.items():
        dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")
    wb.save(path)


def main() -> None:
    HUMAN_V5_DIR.mkdir(parents=True, exist_ok=True)
    LOCKED_V5_DIR.mkdir(parents=True, exist_ok=True)
    dataset = read_csv(DATASET)
    v2 = read_csv(V2_FINAL)
    dev = read_csv(V4_DEV)
    locked = read_csv(V4_LOCKED)
    existing_ids = set(dev["comment_id"]) | set(locked["comment_id"])
    existing_text = set(dev["model_text"].map(normalize_text)) | set(dev["comment_text"].map(normalize_text))
    frame = dataset.merge(
        v2[["comment_id", "probability_negative", "probability_neutral", "probability_positive", "final_sentiment_label"]],
        on="comment_id",
        how="left",
        validate="one_to_one",
    )
    frame = frame.loc[
        ~frame["comment_id"].isin(existing_ids)
        & frame["comment_id"].astype(str).str.strip().ne("")
        & ~frame["comment_id"].astype(str).str.contains("INJ", case=False, regex=False)
    ].copy()
    frame["text_norm"] = frame["text"].map(normalize_text)
    frame = frame.loc[frame["text_norm"].ne("") & ~frame["text_norm"].isin(existing_text)].copy()
    frame = frame.drop_duplicates("comment_id").drop_duplicates("text_norm")
    frame = assign_strata(frame)
    frame["month_bucket"] = frame["timestamp"].map(month_bucket)

    development_plan = [
        ("low_confidence_or_v2_abstained", "low_confidence_sampling", 250),
        ("mixed_sentiment_candidate", "mixed_sentiment_flag_sampling", 150),
        ("question_candidate", "question_flag_sampling", 150),
        ("comparison_brand_candidate", "comparison_brand_flag_sampling", 100),
        ("emoji_or_slang_candidate", "emoji_or_slang_flag_sampling", 100),
        ("short_comment_candidate", "short_comment_flag_sampling", 100),
        ("negative_word_positive_candidate", "negative_word_positive_candidate_sampling", 100),
        ("random_control", "random_control", 50),
    ]
    locked_plan = [
        ("natural_random_observational", "random_control", 250),
        ("low_confidence_or_v2_abstained", "low_confidence_sampling", 150),
        ("question_candidate", "question_flag_sampling", 100),
        ("mixed_sentiment_candidate", "mixed_sentiment_flag_sampling", 75),
        ("comparison_brand_candidate", "comparison_brand_flag_sampling", 75),
        ("short_or_emoji_candidate", "emoji_or_slang_flag_sampling", 50),
    ]
    dev_sample = sample_by_plan(frame, development_plan, "V5D")
    locked_pool = frame.loc[~frame["comment_id"].isin(dev_sample["comment_id"]) & ~frame["text_norm"].isin(dev_sample["text_norm"])].copy()
    locked_sample = sample_by_plan(locked_pool, locked_plan, "V5L")

    dev_candidates = candidate_columns(dev_sample, "DEVELOPMENT_V5_NEW_PENDING")
    locked_candidates = candidate_columns(locked_sample, "LOCKED_TEST_V5_NEW_PENDING")
    dev_candidates.to_csv(HUMAN_V5_DIR / "sentiment_v5_development_candidates.csv", index=False, encoding="utf-8-sig")
    locked_candidates.to_csv(LOCKED_V5_DIR / "sentiment_v5_locked_test_candidates.csv", index=False, encoding="utf-8-sig")

    for annotator in [1, 2]:
        write_workbook(
            HUMAN_V5_DIR / f"sentiment_v5_development_annotator_{annotator}.xlsx",
            annotator_view(dev_candidates),
            f"DEVELOPMENT_ANNOTATOR_{annotator}",
        )
        write_workbook(
            LOCKED_V5_DIR / f"sentiment_v5_locked_test_annotator_{annotator}.xlsx",
            annotator_view(locked_candidates),
            f"LOCKED_TEST_ANNOTATOR_{annotator}",
        )
    adjudication_cols = [
        "annotation_id",
        "comment_id",
        "video_id",
        "product_category",
        "brand_or_video_context",
        "comment_text",
        "parent_comment_text",
        "annotator_1",
        "annotator_2",
        "adjudicated_label",
        "adjudication_note",
    ]
    write_workbook(HUMAN_V5_DIR / "sentiment_v5_development_adjudication.xlsx", dev_candidates[adjudication_cols], "DEVELOPMENT_ADJUDICATION")
    write_workbook(LOCKED_V5_DIR / "sentiment_v5_locked_test_adjudication.xlsx", locked_candidates[adjudication_cols], "LOCKED_TEST_ADJUDICATION")

    dev_audit = dev_sample[
        [
            "annotation_id",
            "comment_id",
            "video_id",
            "product_category",
            "timestamp",
            "month_bucket",
            "sampling_stratum",
            "v2_confidence",
            "v2_margin",
            "final_sentiment_label",
        ]
    ].copy()
    locked_audit = locked_sample[
        [
            "annotation_id",
            "comment_id",
            "video_id",
            "product_category",
            "timestamp",
            "month_bucket",
            "sampling_stratum",
            "v2_confidence",
            "v2_margin",
            "final_sentiment_label",
        ]
    ].copy()
    dev_audit.to_csv(HUMAN_V5_DIR / "development_v5_sampling_internal_audit.csv", index=False, encoding="utf-8-sig")
    locked_audit.to_csv(LOCKED_V5_DIR / "locked_test_v5_sampling_internal_audit.csv", index=False, encoding="utf-8-sig")

    manifest = {
        "status": "SENTIMENT_V5_ANNOTATION_PACKAGES_READY_FOR_HUMAN_LABELING",
        "created_at_utc": utc_now(),
        "development_rows": int(len(dev_candidates)),
        "locked_test_candidate_rows": int(len(locked_candidates)),
        "label_status": "blank_human_labels_required",
        "primary_label_for_rm2_goals": "sentiment_toward_target",
        "existing_development_rows_excluded": int(len(dev)),
        "existing_locked_test_rows_excluded": int(len(locked)),
        "comment_id_overlap_development_locked_v5": int(len(set(dev_candidates["comment_id"]) & set(locked_candidates["comment_id"]))),
        "normalized_text_overlap_development_locked_v5": int(
            len(set(dev_sample["text_norm"].astype(str)) & set(locked_sample["text_norm"].astype(str)))
        ),
        "development_candidate_hash": sha256_dataframe(dev_candidates, ["annotation_id", "comment_id"]),
        "locked_test_candidate_hash": sha256_dataframe(locked_candidates, ["annotation_id", "comment_id"]),
        "source_hashes": {
            "dataset": sha256_file(DATASET),
            "v2_final_observational": sha256_file(V2_FINAL),
            "v4_development_registry": sha256_file(V4_DEV),
            "v4_locked_test_registry": sha256_file(V4_LOCKED),
        },
        "methodology": {
            "locked_test_v4_errors_used_for_sampling": False,
            "locked_test_v4_labels_used_for_sampling_or_training": False,
            "v2_predictions_used_for_sampling_only": True,
            "model_predictions_visible_to_annotators": False,
            "final_labels_auto_filled": False,
            "target_aware_primary_label": "sentiment_toward_target",
        },
        "outputs": {
            "development_candidates": rel(HUMAN_V5_DIR / "sentiment_v5_development_candidates.csv"),
            "locked_candidates": rel(LOCKED_V5_DIR / "sentiment_v5_locked_test_candidates.csv"),
            "development_annotator_1": rel(HUMAN_V5_DIR / "sentiment_v5_development_annotator_1.xlsx"),
            "development_annotator_2": rel(HUMAN_V5_DIR / "sentiment_v5_development_annotator_2.xlsx"),
            "locked_annotator_1": rel(LOCKED_V5_DIR / "sentiment_v5_locked_test_annotator_1.xlsx"),
            "locked_annotator_2": rel(LOCKED_V5_DIR / "sentiment_v5_locked_test_annotator_2.xlsx"),
        },
    }
    write_json(HUMAN_V5_DIR / "SENTIMENT_V5_DEVELOPMENT_PACKAGE_MANIFEST.json", manifest)
    locked_freeze_manifest = {
        **manifest,
        "status": "LOCKED_TEST_V5_CANDIDATE_LIST_FROZEN_PENDING_HUMAN_LABELS",
        "freeze_scope": "comment_id_list_and_sampling_manifest_only",
        "final_label_freeze_status": "BLOCKED_PENDING_TWO_ANNOTATORS_AND_ADJUDICATION",
        "final_label_dataset_hash": "",
        "annotation_guideline_version": "SENTIMENT_TARGET_AWARE_ANNOTATION_GUIDE_V5",
        "candidate_comment_id_hash": sha256_dataframe(locked_candidates, ["comment_id"]),
    }
    write_json(LOCKED_V5_DIR / "LOCKED_TEST_V5_FREEZE_MANIFEST.json", locked_freeze_manifest)
    print(json.dumps({"development_rows": len(dev_candidates), "locked_test_candidate_rows": len(locked_candidates)}, indent=2), flush=True)


if __name__ == "__main__":
    main()
