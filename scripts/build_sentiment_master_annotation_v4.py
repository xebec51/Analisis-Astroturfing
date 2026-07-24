from __future__ import annotations

import hashlib
import html
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output/rm2_sentiment/validation/human_master_v4"

DATASET = ROOT / "dataset.csv"
OBS_V2 = ROOT / "output/rm2_sentiment/final/comment_sentiment_v2_observational.csv"
PREVIOUS_REGISTRY = ROOT / "output/rm2_sentiment/validation/human_v2_positive_recall/human_label_registry.csv"
ACTIVE_LEARNING_BLIND = ROOT / "output/rm2_sentiment/validation/human_v2_positive_recall/positive_active_learning_blind.csv"
ACTIVE_LEARNING_MANIFEST = ROOT / "output/rm2_sentiment/validation/human_v2_positive_recall/positive_active_learning_sampling_manifest.csv"
LOCKED_TEMPLATE = ROOT / "output/rm2_sentiment/validation/human_v2_positive_recall/new_locked_test_final.csv"
LOCKED_SAMPLING_MANIFEST = ROOT / "output/rm2_sentiment/validation/human_v2_positive_recall/new_locked_test_sampling_manifest.csv"
EXACT_DUP_MEMBERS = ROOT / "output/rm2_comment_similarity/exact_duplicate_comment_group_members.csv"
NEAR_DUP_MEMBERS = ROOT / "output/rm2_comment_similarity/near_similar_comment_cluster_members.csv"

SOURCE_DIRS = [
    ROOT / "output/rm2_sentiment/validation/human_v1",
    ROOT / "output/rm2_sentiment/validation/human_v2",
    ROOT / "output/rm2_sentiment/validation/human_v3",
    ROOT / "output/rm2_sentiment/validation/human_v2_positive_recall",
]

LABELS = ["Negative", "Neutral", "Positive", "Uncertain", "No Text"]
EVALUABLE = ["Negative", "Neutral", "Positive"]
ROLE_VALUES = [
    "HISTORICAL_DEVELOPMENT_FINAL",
    "DEVELOPMENT_NEW_PENDING",
    "DEVELOPMENT_NEW_FINAL",
    "LEGACY_TEST_PROVENANCE",
    "LOCKED_TEST_NEW_PENDING",
    "LOCKED_TEST_NEW_FINAL",
    "EXCLUDED",
]
PENDING_DEV_TARGET = 1300
PENDING_LOCKED_TARGET = 700
ADDITIONAL_DEV_TARGET = 800
ADDITIONAL_LOCKED_TARGET = 100

MASTER_COLUMNS = [
    "master_annotation_id",
    "comment_id",
    "video_id",
    "comment_text",
    "model_text",
    "product_category",
    "brand_or_video_context",
    "source_file",
    "source_version",
    "source_role",
    "annotation_role",
    "split_lock",
    "historical_final_label",
    "annotator_1_label",
    "annotator_2_label",
    "adjudicated_label",
    "final_human_label",
    "annotation_status",
    "disagreement_flag",
    "adjudication_required",
    "evaluable_three_class",
    "exclusion_reason",
    "text_cluster_id",
    "exact_duplicate_group_id",
    "near_duplicate_cluster_id",
    "sampling_stratum",
    "sampling_batch",
    "created_at",
    "notes",
]
ANNOTATOR_COLUMNS = [
    "annotation_id",
    "comment_id",
    "video_id",
    "product_category",
    "brand_or_video_context",
    "comment_text",
    "human_label",
    "confidence_annotation",
    "needs_context",
    "annotator_notes",
]
BANNED_ANNOTATOR_TERMS = [
    "prediction",
    "probability",
    "threshold",
    "hcc",
    "actor",
    "goal",
    "disagreement",
    "model",
]


def read_csv(path: Path, required: bool = True) -> pd.DataFrame:
    if not path.exists():
        if required:
            raise FileNotFoundError(path)
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, keep_default_na=False, low_memory=False)


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def sha256_file(path: Path) -> str:
    if not path.exists():
        return ""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_blank(value: object) -> str:
    if isinstance(value, pd.Series):
        for item in value.tolist():
            text = norm_blank(item)
            if text:
                return text
        return ""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null", "<na>"}:
        return ""
    return text


def canonical_label(value: object) -> str:
    text = norm_blank(value)
    mapping = {label.casefold(): label for label in LABELS}
    return mapping.get(text.casefold(), text)


def normalize_text(text: object) -> str:
    s = norm_blank(text)
    s = unicodedata.normalize("NFKC", s)
    s = html.unescape(s)
    s = "".join(ch for ch in s if unicodedata.category(ch)[0] != "C" or ch in "\t\n\r")
    return re.sub(r"\s+", " ", s).strip()


def normalized_group_text(text: object) -> str:
    return normalize_text(text).casefold()


def model_text(text: object) -> str:
    s = normalize_text(text)
    s = re.sub(r"https?://\S+|www\.\S+", "HTTPURL", s, flags=re.IGNORECASE)
    s = re.sub(r"(?<!\w)@\w+", "@USER", s)
    return re.sub(r"\s+", " ", s).strip()


def stable_hash(text: str, n: int = 16) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:n]


def is_inj(comment_id: object) -> bool:
    text = norm_blank(comment_id).upper()
    return text.startswith("INJ") or "SYNTHETIC" in text


def no_text(text: object) -> bool:
    s = norm_blank(text)
    return s.casefold() in {"", "nan", "none", "null", "<na>", "[deleted]", "deleted"}


def discover_annotation_sources() -> list[Path]:
    patterns = re.compile(
        r"(annotation|annotator|adjudication|validated|locked_test|active_learning|final|human|label)",
        flags=re.IGNORECASE,
    )
    paths: list[Path] = []
    for directory in SOURCE_DIRS:
        if not directory.exists():
            continue
        for path in directory.rglob("*"):
            if path.is_file() and path.suffix.lower() in {".csv", ".xlsx", ".xls", ".json", ".md"} and patterns.search(path.name):
                paths.append(path)
    return sorted(set(paths), key=lambda p: rel(p))


def label_columns(columns: list[str]) -> list[str]:
    return [
        col
        for col in columns
        if re.search(r"(final|adjudicated|human|annotator).*label|label", col, flags=re.IGNORECASE)
    ]


def source_status(path: Path, columns: list[str], frame: pd.DataFrame) -> tuple[str, str, bool, bool, bool]:
    name = path.name.lower()
    rel_path = rel(path).lower()
    status = "provenance"
    role = "provenance"
    opened_test = "locked_test" in name or "final_test" in name or "final_test" in rel_path
    can_enter_dev = False
    provenance_only = True
    if any(token in name for token in ["blind", "annotator_1", "annotator_2"]) and "positive_recall" in rel_path:
        status = "pending"
    if "adjudication" in name and "positive_recall" in rel_path:
        status = "pending"
    if any(token in name for token in ["validated", "adjudication_template_final", "replacement_adjudication_final", "human_label_registry"]):
        status = "final_or_registry"
    if "locked_test" in name or "final_test" in name:
        role = "test_or_locked"
    elif "development" in name or "active_learning" in name:
        role = "development_or_pending"
    if path == PREVIOUS_REGISTRY:
        status = "consolidated_registry"
        role = "mixed"
    if "development" in frame.columns and frame["development"].astype(str).str.len().gt(0).any():
        role = "development_or_pending"
    if status in {"final_or_registry", "consolidated_registry"} and role != "test_or_locked":
        can_enter_dev = True
        provenance_only = False
    if opened_test:
        can_enter_dev = False
        provenance_only = True
    return status, role, opened_test, can_enter_dev, provenance_only


def inventory_sources(paths: list[Path]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    inventory_rows = []
    checksum_rows = []
    schema_rows = []
    count_rows = []
    for path in paths:
        checksum_rows.append({"path": rel(path), "sha256": sha256_file(path), "bytes": path.stat().st_size})
        if path.suffix.lower() != ".csv":
            inventory_rows.append(
                {
                    "path": rel(path),
                    "rows": "",
                    "unique_comment_id": "",
                    "status": "non_csv_provenance",
                    "role": "provenance",
                    "opened_as_test": "",
                    "can_enter_development": False,
                    "provenance_only": True,
                    "human_label_columns": "",
                    "adjudication_columns": "",
                }
            )
            schema_rows.append({"path": rel(path), "columns": "", "read_status": "NON_CSV_SKIPPED"})
            continue
        try:
            frame = read_csv(path)
            labels = label_columns(frame.columns.tolist())
            adjudication_cols = [col for col in frame.columns if "adjudicat" in col.lower()]
            status, role, opened, can_dev, provenance_only = source_status(path, frame.columns.tolist(), frame)
            comment_unique = frame["comment_id"].nunique() if "comment_id" in frame.columns else ""
            inventory_rows.append(
                {
                    "path": rel(path),
                    "rows": len(frame),
                    "unique_comment_id": comment_unique,
                    "status": status,
                    "role": role,
                    "opened_as_test": opened,
                    "can_enter_development": can_dev,
                    "provenance_only": provenance_only,
                    "human_label_columns": ";".join(labels),
                    "adjudication_columns": ";".join(adjudication_cols),
                }
            )
            schema_rows.append({"path": rel(path), "columns": "|".join(frame.columns), "read_status": "OK"})
            count = {"path": rel(path), "rows": len(frame), "unique_comment_id": comment_unique}
            for col in labels[:5]:
                values = frame[col].map(canonical_label)
                for label in LABELS + [""]:
                    count[f"{col}:{label or 'blank'}"] = int(values.eq(label).sum())
            count_rows.append(count)
        except Exception as exc:  # pragma: no cover - inventory should keep going.
            inventory_rows.append(
                {
                    "path": rel(path),
                    "rows": "",
                    "unique_comment_id": "",
                    "status": f"read_error:{exc}",
                    "role": "unknown",
                    "opened_as_test": "",
                    "can_enter_development": False,
                    "provenance_only": True,
                    "human_label_columns": "",
                    "adjudication_columns": "",
                }
            )
            schema_rows.append({"path": rel(path), "columns": "", "read_status": f"ERROR:{exc}"})
    return pd.DataFrame(inventory_rows), pd.DataFrame(checksum_rows), pd.DataFrame(schema_rows), pd.DataFrame(count_rows)


def external_cluster_maps() -> tuple[dict[str, str], dict[str, str]]:
    exact_map: dict[str, str] = {}
    near_map: dict[str, str] = {}
    exact = read_csv(EXACT_DUP_MEMBERS, required=False)
    if not exact.empty and {"comment_id", "exact_group_id"}.issubset(exact.columns):
        exact_map.update(dict(zip(exact["comment_id"].map(norm_blank), exact["exact_group_id"].map(norm_blank))))
    near = read_csv(NEAR_DUP_MEMBERS, required=False)
    if not near.empty and {"comment_id", "cluster_id"}.issubset(near.columns):
        near_map.update(dict(zip(near["comment_id"].map(norm_blank), near["cluster_id"].map(norm_blank))))
    return exact_map, near_map


def add_cluster_columns(frame: pd.DataFrame, exact_map: dict[str, str], near_map: dict[str, str]) -> pd.DataFrame:
    out = frame.copy()
    normalized = out["comment_text"].map(normalized_group_text)
    exact_hash = normalized.map(lambda x: "EXACT_TEXT_" + stable_hash(x))
    out["exact_duplicate_group_id"] = out["comment_id"].map(exact_map).fillna(exact_hash)
    out["near_duplicate_cluster_id"] = out["comment_id"].map(near_map).fillna("")
    out["text_cluster_id"] = out["exact_duplicate_group_id"]
    return out


def base_master_row(row: pd.Series, role: str, batch: str, stratum: str = "", notes: str = "") -> dict[str, object]:
    comment_text = norm_blank(row.get("comment_text", row.get("comment_text_original", row.get("text", ""))))
    if not comment_text:
        comment_text = norm_blank(row.get("comment_text_original.1", ""))
    final_label = canonical_label(row.get("final_human_label", row.get("final_sentiment_label", "")))
    historical_label = canonical_label(row.get("historical_final_label", final_label if role in {"HISTORICAL_DEVELOPMENT_FINAL", "LEGACY_TEST_PROVENANCE", "EXCLUDED"} else ""))
    split_lock = {
        "HISTORICAL_DEVELOPMENT_FINAL": "DEVELOPMENT_LOCKED_HISTORICAL",
        "DEVELOPMENT_NEW_PENDING": "DEVELOPMENT_PENDING_LOCKED",
        "DEVELOPMENT_NEW_FINAL": "DEVELOPMENT_FINAL_LOCKED",
        "LEGACY_TEST_PROVENANCE": "LEGACY_TEST_READONLY",
        "LOCKED_TEST_NEW_PENDING": "LOCKED_TEST_PENDING_LOCKED",
        "LOCKED_TEST_NEW_FINAL": "LOCKED_TEST_FINAL_LOCKED",
        "EXCLUDED": "EXCLUDED_READONLY",
    }[role]
    pending = role in {"DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"}
    excluded_reason = norm_blank(row.get("exclusion_reason", ""))
    if is_inj(row.get("comment_id", "")):
        excluded_reason = "injected_or_synthetic_comment_id"
    elif no_text(comment_text) and role != "EXCLUDED":
        excluded_reason = "no_readable_text"
    return {
        "master_annotation_id": "",
        "comment_id": norm_blank(row.get("comment_id", "")),
        "video_id": norm_blank(row.get("video_id", "")),
        "comment_text": comment_text,
        "model_text": model_text(comment_text),
        "product_category": norm_blank(row.get("product_category", "")),
        "brand_or_video_context": norm_blank(row.get("brand_or_video_context", row.get("product_category", ""))),
        "source_file": norm_blank(row.get("source_file", "")),
        "source_version": norm_blank(row.get("source_version", "")),
        "source_role": norm_blank(row.get("source_role", row.get("source_sample_role", row.get("sample_role", "")))),
        "annotation_role": role,
        "split_lock": split_lock,
        "historical_final_label": "" if pending else historical_label,
        "annotator_1_label": "" if pending else canonical_label(row.get("annotator_1_label", "")),
        "annotator_2_label": "" if pending else canonical_label(row.get("annotator_2_label", "")),
        "adjudicated_label": "" if pending else canonical_label(row.get("adjudicated_label", "")),
        "final_human_label": "" if pending else final_label,
        "annotation_status": "PENDING_HUMAN_LABELING" if pending else norm_blank(row.get("annotation_status", "FINAL_HUMAN")),
        "disagreement_flag": False,
        "adjudication_required": False if not pending else True,
        "evaluable_three_class": (final_label in EVALUABLE) if not pending else "",
        "exclusion_reason": excluded_reason,
        "text_cluster_id": "",
        "exact_duplicate_group_id": "",
        "near_duplicate_cluster_id": "",
        "sampling_stratum": stratum,
        "sampling_batch": batch,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "notes": notes,
    }


def historical_master_rows() -> pd.DataFrame:
    reg = read_csv(PREVIOUS_REGISTRY)
    rows = []
    for _, row in reg.iterrows():
        split = norm_blank(row.get("split_family", ""))
        registry_role = norm_blank(row.get("registry_role", ""))
        label = canonical_label(row.get("final_sentiment_label", ""))
        if split == "development" and label in EVALUABLE:
            role = "HISTORICAL_DEVELOPMENT_FINAL"
            status = "FINAL_HISTORICAL_DEVELOPMENT"
            reason = ""
        elif split == "legacy_diagnostic_test_already_opened":
            role = "LEGACY_TEST_PROVENANCE"
            status = "READONLY_LEGACY_TEST_PROVENANCE"
            reason = ""
        else:
            role = "EXCLUDED"
            status = "EXCLUDED_READONLY"
            reason = registry_role if registry_role.startswith("excluded") else "non_evaluable_or_not_allowed"
        source = row.copy()
        source["comment_text"] = row.get("comment_text_original", "")
        source["historical_final_label"] = label
        source["final_human_label"] = label
        source["adjudicated_label"] = label
        source["annotation_status"] = status
        source["exclusion_reason"] = reason
        rows.append(base_master_row(source, role, "historical_v1_v2_v3", notes="Consolidated from previous human positive-recall registry."))
    return pd.DataFrame(rows)


def clean_pending_frame(frame: pd.DataFrame, source_file: Path, source_version: str, role: str, batch: str) -> pd.DataFrame:
    out = frame.copy()
    out = out.loc[:, ~out.columns.duplicated()].copy()
    if "comment_text_original" not in out.columns and "text" in out.columns:
        out["comment_text_original"] = out["text"]
    if "brand_or_video_context" not in out.columns:
        out["brand_or_video_context"] = out.get("product_category", "")
    out["source_file"] = rel(source_file)
    out["source_version"] = source_version
    out["source_role"] = out.get("sample_role", role.lower())
    rows = []
    for _, row in out.iterrows():
        rows.append(base_master_row(row, role, batch))
    return pd.DataFrame(rows)


def score_columns(obs: pd.DataFrame) -> pd.DataFrame:
    out = obs.copy()
    for col in ["probability_negative", "probability_neutral", "probability_positive", "max_probability"]:
        out[col] = pd.to_numeric(out.get(col, 0.0), errors="coerce").fillna(0.0)
    probs = out[["probability_negative", "probability_neutral", "probability_positive"]].to_numpy()
    sorted_probs = np.sort(probs, axis=1)
    out["top_margin"] = sorted_probs[:, -1] - sorted_probs[:, -2]
    lower = out["text"].fillna("").astype(str).str.casefold()
    out["flag_question"] = lower.str.contains(r"\?|apa|berapa|gimana|bagaimana|aman|cocok|boleh", regex=True)
    out["flag_positive_hard"] = lower.str.contains(r"hasil|cocok|worth|rekomen|wajib|mantul|glow|cerah|pudar|suka|love", regex=True)
    out["flag_negative_hard"] = lower.str.contains(r"jerawat|gatal|perih|kecewa|bohong|palsu|buruk|parah|rusak|takut|nggak cocok|gak cocok|ga cocok", regex=True)
    out["flag_neutral_hard"] = lower.str.contains(r"harga|beli dimana|spill|link|cara pakai|berapa|mau coba|belum coba|tag|kak", regex=True)
    out["flag_low_confidence"] = (out["max_probability"] < 0.55) | (out["top_margin"] <= 0.15) | out.get("abstained", "").astype(str).str.lower().eq("true")
    out["positive_score"] = out["probability_positive"] + out["flag_positive_hard"].astype(int) * 0.35 + out["flag_low_confidence"].astype(int) * 0.10
    out["negative_score"] = out["probability_negative"] + out["flag_negative_hard"].astype(int) * 0.40 + out["flag_question"].astype(int) * 0.05
    out["neutral_score"] = out["probability_neutral"] + out["flag_neutral_hard"].astype(int) * 0.35 + out["flag_question"].astype(int) * 0.10
    out["lowconf_score"] = (1 - out["max_probability"]) + (0.5 - out["top_margin"]).clip(lower=0) + out["flag_low_confidence"].astype(int) * 0.5
    out["random_score"] = out["comment_id"].map(lambda x: int(stable_hash(str(x), 8), 16))
    return out


def obs_candidates(exclude_ids: set[str], exclude_clusters: set[str], exact_map: dict[str, str], near_map: dict[str, str]) -> pd.DataFrame:
    obs = score_columns(read_csv(OBS_V2))
    obs = obs.loc[~obs["comment_id"].map(norm_blank).isin(exclude_ids)].copy()
    obs = obs.loc[~obs["comment_id"].map(is_inj)].copy()
    obs = obs.loc[~obs["text"].map(no_text)].copy()
    temp = pd.DataFrame(
        {
            "comment_id": obs["comment_id"].map(norm_blank),
            "comment_text": obs["text"].map(norm_blank),
        }
    )
    temp = add_cluster_columns(temp, exact_map, near_map)
    obs = obs.merge(temp[["comment_id", "text_cluster_id", "exact_duplicate_group_id", "near_duplicate_cluster_id"]], on="comment_id", how="left")
    if exclude_clusters:
        obs = obs.loc[~obs["text_cluster_id"].isin(exclude_clusters)].copy()
    return obs.reset_index(drop=True)


def select_by_strata(pool: pd.DataFrame, targets: list[tuple[str, int, str]], used_ids: set[str], used_clusters: set[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    selected_parts = []
    audit_rows = []
    for stratum, target, score_col in targets:
        available = pool.loc[~pool["comment_id"].isin(used_ids) & ~pool["text_cluster_id"].isin(used_clusters)].copy()
        if score_col != "random_score":
            available = available.sort_values([score_col, "top_margin", "comment_id"], ascending=[False, True, True])
        else:
            available = available.sort_values([score_col, "comment_id"])
        chosen = available.head(target).copy()
        chosen["master_sampling_stratum"] = stratum
        selected_parts.append(chosen)
        used_ids.update(chosen["comment_id"].map(norm_blank))
        used_clusters.update(chosen["text_cluster_id"].map(norm_blank))
        audit_rows.append({"sampling_stratum": stratum, "target": target, "selected": len(chosen), "score_column": score_col})
    selected = pd.concat(selected_parts, ignore_index=True) if selected_parts else pd.DataFrame()
    return selected, pd.DataFrame(audit_rows)


def assign_locked_strata(locked: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    frame = locked.copy()
    if "text" not in frame.columns:
        frame["text"] = ""
    if "comment_text_original" in frame.columns:
        frame["text"] = np.where(frame["text"].map(norm_blank).ne(""), frame["text"], frame["comment_text_original"])
    frame = score_columns(frame)
    used_ids: set[str] = set()
    used_clusters: set[str] = set()
    targets = [
        ("locked_disagreement_low_confidence", 100, "lowconf_score"),
        ("locked_positive_enriched", 150, "positive_score"),
        ("locked_negative_enriched", 100, "negative_score"),
        ("locked_neutral_enriched", 100, "neutral_score"),
        ("locked_natural_random_observational", 250, "random_score"),
    ]
    selected, audit = select_by_strata(frame, targets, used_ids, used_clusters)
    if len(selected) != PENDING_LOCKED_TARGET:
        raise AssertionError(f"Locked pending selection expected {PENDING_LOCKED_TARGET}, found {len(selected)}")
    return selected, audit


def build_pending_rows(historical: pd.DataFrame, exact_map: dict[str, str], near_map: dict[str, str]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    active = clean_pending_frame(read_csv(ACTIVE_LEARNING_BLIND), ACTIVE_LEARNING_BLIND, "positive_active_learning_v2_positive_recall", "DEVELOPMENT_NEW_PENDING", "pending_existing_active_learning_500")
    locked_existing = clean_pending_frame(read_csv(LOCKED_TEMPLATE), LOCKED_TEMPLATE, "locked_test_template_v2_positive_recall", "LOCKED_TEST_NEW_PENDING", "pending_existing_locked_test_600")

    active = add_cluster_columns(active, exact_map, near_map)
    locked_existing = add_cluster_columns(locked_existing, exact_map, near_map)
    historical = add_cluster_columns(historical, exact_map, near_map)

    historical_dev_clusters = set(historical.loc[historical["annotation_role"].eq("HISTORICAL_DEVELOPMENT_FINAL"), "text_cluster_id"])
    all_existing_ids = set(historical["comment_id"]) | set(active["comment_id"]) | set(locked_existing["comment_id"])

    locked_pool = read_csv(LOCKED_SAMPLING_MANIFEST).rename(columns={"text": "comment_text_original"})
    locked_pool = locked_pool.merge(
        locked_existing[["comment_id", "text_cluster_id", "exact_duplicate_group_id", "near_duplicate_cluster_id"]],
        on="comment_id",
        how="right",
    )
    locked_existing_for_strata = locked_pool.copy()

    locked_add_pool = obs_candidates(
        exclude_ids=all_existing_ids,
        exclude_clusters=historical_dev_clusters | set(active["text_cluster_id"]) | set(locked_existing["text_cluster_id"]),
        exact_map=exact_map,
        near_map=near_map,
    )
    locked_add_targets = [("locked_additional_candidate_pool", ADDITIONAL_LOCKED_TARGET, "lowconf_score")]
    locked_add, locked_add_audit = select_by_strata(locked_add_pool, locked_add_targets, set(), set())
    if len(locked_add) != ADDITIONAL_LOCKED_TARGET:
        raise AssertionError("Could not select 100 additional locked-test candidates.")
    locked_add["locked_test_item_id"] = [f"V4LTADD{i:04d}" for i in range(1, len(locked_add) + 1)]
    locked_add["sample_role"] = "locked_test_v4_additional_candidate"
    locked_add_rows = clean_pending_frame(
        locked_add.rename(columns={"text": "comment_text_original"}),
        LOCKED_TEMPLATE,
        "locked_test_v4_additional_sampling",
        "LOCKED_TEST_NEW_PENDING",
        "pending_locked_test_additional_100",
    )
    locked_add_rows["sampling_stratum"] = "locked_additional_candidate_pool"
    locked_add_rows = add_cluster_columns(locked_add_rows, exact_map, near_map)

    locked_existing_strata_input = locked_existing_for_strata.rename(columns={"text": "comment_text_original"})
    locked_add_strata_input = locked_add.rename(columns={"text": "comment_text_original"})
    locked_existing_strata_input = locked_existing_strata_input.loc[:, ~locked_existing_strata_input.columns.duplicated()].copy()
    locked_add_strata_input = locked_add_strata_input.loc[:, ~locked_add_strata_input.columns.duplicated()].copy()
    locked_combined_for_strata = pd.concat(
        [
            locked_existing_strata_input,
            locked_add_strata_input,
        ],
        ignore_index=True,
        sort=False,
    )
    for col in ["text_cluster_id", "exact_duplicate_group_id", "near_duplicate_cluster_id"]:
        if col not in locked_combined_for_strata.columns:
            locked_combined_for_strata[col] = ""
    locked_strata, locked_strata_audit = assign_locked_strata(locked_combined_for_strata)
    locked_stratum_map = dict(zip(locked_strata["comment_id"], locked_strata["master_sampling_stratum"]))
    locked_existing["sampling_stratum"] = locked_existing["comment_id"].map(locked_stratum_map).fillna("locked_existing_audit_remainder")
    locked_add_rows["sampling_stratum"] = locked_add_rows["comment_id"].map(locked_stratum_map).fillna("locked_additional_audit_remainder")
    locked_pending = pd.concat([locked_existing, locked_add_rows], ignore_index=True)
    if len(locked_pending) != PENDING_LOCKED_TARGET:
        raise AssertionError(f"Locked pending expected 700, found {len(locked_pending)}")

    locked_clusters = set(locked_pending["text_cluster_id"])
    active_conflict = active["text_cluster_id"].isin(locked_clusters)
    if active_conflict.any():
        active.loc[active_conflict, "annotation_role"] = "EXCLUDED"
        active.loc[active_conflict, "split_lock"] = "EXCLUDED_READONLY"
        active.loc[active_conflict, "annotation_status"] = "EXCLUDED_DUPLICATE_WITH_LOCKED_TEST"
        active.loc[active_conflict, "exclusion_reason"] = "duplicate_cluster_with_locked_test_pending"

    active_kept = active.loc[active["annotation_role"].eq("DEVELOPMENT_NEW_PENDING")].copy()
    active_kept = active_kept.sort_values("comment_id").reset_index(drop=True)
    active_kept["sampling_stratum"] = ["development_positive_enriched" if i < 400 else "development_disagreement_low_confidence" for i in range(len(active_kept))]

    need_dev_additional = PENDING_DEV_TARGET - len(active_kept)
    if need_dev_additional < ADDITIONAL_DEV_TARGET:
        need_dev_additional = ADDITIONAL_DEV_TARGET
    exclude_ids = all_existing_ids | set(locked_add["comment_id"])
    dev_pool = obs_candidates(exclude_ids=exclude_ids, exclude_clusters=locked_clusters, exact_map=exact_map, near_map=near_map)
    dev_targets = [
        ("development_negative_enriched", 300, "negative_score"),
        ("development_neutral_enriched", 300, "neutral_score"),
        ("development_disagreement_low_confidence", 100, "lowconf_score"),
        ("development_random_control", 100, "random_score"),
    ]
    dev_add, dev_add_audit = select_by_strata(dev_pool, dev_targets, set(), set(locked_clusters))
    remaining_needed = PENDING_DEV_TARGET - len(active_kept) - len(dev_add)
    if remaining_needed > 0:
        filler, filler_audit = select_by_strata(
            dev_pool,
            [("development_positive_enriched", remaining_needed, "positive_score")],
            set(dev_add["comment_id"]),
            set(dev_add["text_cluster_id"]) | locked_clusters,
        )
        dev_add = pd.concat([dev_add, filler], ignore_index=True)
        dev_add_audit = pd.concat([dev_add_audit, filler_audit], ignore_index=True)
    if len(dev_add) != PENDING_DEV_TARGET - len(active_kept):
        raise AssertionError(f"Development additional selection mismatch: {len(dev_add)}")
    dev_add["annotation_item_id"] = [f"V4DEVADD{i:04d}" for i in range(1, len(dev_add) + 1)]
    dev_add["sample_role"] = "development_v4_additional_candidate"
    dev_add_rows = clean_pending_frame(
        dev_add.rename(columns={"text": "comment_text_original"}),
        ACTIVE_LEARNING_BLIND,
        "development_v4_additional_sampling",
        "DEVELOPMENT_NEW_PENDING",
        "pending_development_additional_800",
    )
    dev_add_rows["sampling_stratum"] = dev_add["master_sampling_stratum"].to_numpy()
    dev_add_rows = add_cluster_columns(dev_add_rows, exact_map, near_map)
    dev_pending = pd.concat([active_kept, dev_add_rows], ignore_index=True)
    if len(dev_pending) != PENDING_DEV_TARGET:
        raise AssertionError(f"Development pending expected 1300, found {len(dev_pending)}")

    internal_dev = pd.concat(
        [
            read_csv(ACTIVE_LEARNING_MANIFEST).assign(master_sampling_stratum=active_kept["sampling_stratum"].reindex(range(len(read_csv(ACTIVE_LEARNING_MANIFEST)))).fillna("development_active_learning_existing")),
            dev_add.assign(master_sampling_stratum=dev_add["master_sampling_stratum"]),
        ],
        ignore_index=True,
        sort=False,
    )
    internal_locked = pd.concat(
        [
            read_csv(LOCKED_SAMPLING_MANIFEST).assign(master_sampling_stratum=lambda df: df["comment_id"].map(locked_stratum_map).fillna("locked_existing_unassigned")),
            locked_add.assign(master_sampling_stratum=lambda df: df["comment_id"].map(locked_stratum_map).fillna("locked_additional_unassigned")),
        ],
        ignore_index=True,
        sort=False,
    )
    sampling_audit = pd.concat(
        [
            locked_add_audit.assign(package="locked_additional"),
            locked_strata_audit.assign(package="locked_final_700"),
            dev_add_audit.assign(package="development_additional"),
        ],
        ignore_index=True,
        sort=False,
    )
    excluded_active = active.loc[active["annotation_role"].eq("EXCLUDED")].copy()
    pending = pd.concat([dev_pending, locked_pending, excluded_active], ignore_index=True, sort=False)
    return pending, internal_dev, internal_locked, sampling_audit


def finalize_master(master: pd.DataFrame) -> pd.DataFrame:
    out = master.copy()
    out = out.drop_duplicates("comment_id", keep="first").reset_index(drop=True)
    out["master_annotation_id"] = [f"V4MA{i:05d}" for i in range(1, len(out) + 1)]
    for col in MASTER_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    out["disagreement_flag"] = out["disagreement_flag"].fillna(False).astype(bool)
    out["adjudication_required"] = out["adjudication_required"].fillna(False).astype(bool)
    return out[MASTER_COLUMNS]


def conflict_and_leakage_audits(master: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    duplicate_rows = []
    for comment_id, group in master.groupby("comment_id"):
        if len(group) > 1:
            duplicate_rows.append(
                {
                    "comment_id": comment_id,
                    "n_rows": len(group),
                    "roles": ";".join(sorted(set(group["annotation_role"]))),
                    "resolution": "first_role_by_locked_priority_retained",
                }
            )
    duplicate_resolution = pd.DataFrame(duplicate_rows or [{"comment_id": "", "n_rows": 0, "roles": "", "resolution": "no_duplicate_comment_id"}])

    role_by_comment = master.groupby("comment_id")["annotation_role"].nunique().reset_index(name="n_roles")
    comment_conflicts = role_by_comment.loc[role_by_comment["n_roles"].gt(1)].copy()
    if comment_conflicts.empty:
        comment_conflicts = pd.DataFrame([{"comment_id": "", "n_roles": 0, "status": "PASS", "notes": "No comment_id role conflicts after deduplication."}])
    else:
        comment_conflicts["status"] = "FAIL"
        comment_conflicts["notes"] = "Comment_id appears with multiple roles."

    text_conflict_rows = []
    for cluster, group in master.groupby("text_cluster_id"):
        roles = set(group["annotation_role"])
        dev = bool(roles & {"HISTORICAL_DEVELOPMENT_FINAL", "DEVELOPMENT_NEW_PENDING", "DEVELOPMENT_NEW_FINAL"})
        test = bool(roles & {"LOCKED_TEST_NEW_PENDING", "LOCKED_TEST_NEW_FINAL"})
        if dev and test:
            text_conflict_rows.append(
                {
                    "text_cluster_id": cluster,
                    "n_rows": len(group),
                    "roles": ";".join(sorted(roles)),
                    "comment_ids": "|".join(group["comment_id"].astype(str).head(50)),
                    "status": "FAIL",
                    "notes": "Hard text cluster crosses development and locked test.",
                }
            )
    text_conflicts = pd.DataFrame(text_conflict_rows or [{"text_cluster_id": "", "n_rows": 0, "roles": "", "comment_ids": "", "status": "PASS", "notes": "No hard text cluster leakage."}])

    leakage_rows = []
    dev_ids = set(master.loc[master["annotation_role"].isin(["HISTORICAL_DEVELOPMENT_FINAL", "DEVELOPMENT_NEW_PENDING", "DEVELOPMENT_NEW_FINAL"]), "comment_id"])
    test_ids = set(master.loc[master["annotation_role"].isin(["LOCKED_TEST_NEW_PENDING", "LOCKED_TEST_NEW_FINAL"]), "comment_id"])
    leakage_rows.append({"audit_type": "comment_id_development_locked_overlap", "observed": len(dev_ids & test_ids), "status": "PASS" if not (dev_ids & test_ids) else "FAIL"})
    leakage_rows.append({"audit_type": "text_cluster_development_locked_overlap", "observed": int(text_conflicts["status"].eq("FAIL").sum()), "status": "PASS" if not text_conflicts["status"].eq("FAIL").any() else "FAIL"})
    dev_videos = set(master.loc[master["annotation_role"].isin(["HISTORICAL_DEVELOPMENT_FINAL", "DEVELOPMENT_NEW_PENDING"]), "video_id"]) - {""}
    test_videos = set(master.loc[master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING"), "video_id"]) - {""}
    leakage_rows.append({"audit_type": "video_id_development_locked_overlap_soft", "observed": len(dev_videos & test_videos), "status": "WARN" if dev_videos & test_videos else "PASS"})
    split_leakage = pd.DataFrame(leakage_rows)

    exclusion = master.loc[master["annotation_role"].eq("EXCLUDED")].groupby("exclusion_reason", dropna=False).size().rename("count").reset_index()
    if exclusion.empty:
        exclusion = pd.DataFrame([{"exclusion_reason": "", "count": 0}])
    return duplicate_resolution, split_leakage, comment_conflicts, text_conflicts, exclusion


def annotator_view(master: pd.DataFrame, role: str) -> pd.DataFrame:
    rows = master.loc[master["annotation_role"].eq(role)].copy()
    view = pd.DataFrame(
        {
            "annotation_id": rows["master_annotation_id"],
            "comment_id": rows["comment_id"],
            "video_id": rows["video_id"],
            "product_category": rows["product_category"],
            "brand_or_video_context": rows["brand_or_video_context"],
            "comment_text": rows["comment_text"],
            "human_label": "",
            "confidence_annotation": "",
            "needs_context": "",
            "annotator_notes": "",
        }
    )
    return view[ANNOTATOR_COLUMNS]


def adjudication_view(master: pd.DataFrame, role: str) -> pd.DataFrame:
    base = annotator_view(master, role)
    return base.assign(
        annotator_1_label="",
        annotator_1_notes="",
        annotator_2_label="",
        annotator_2_notes="",
        adjudicated_label="",
        adjudication_notes="",
    )


def historical_view(master: pd.DataFrame) -> pd.DataFrame:
    keep = master.loc[master["annotation_role"].isin(["HISTORICAL_DEVELOPMENT_FINAL", "LEGACY_TEST_PROVENANCE", "EXCLUDED"])].copy()
    return keep[
        [
            "master_annotation_id",
            "comment_id",
            "video_id",
            "product_category",
            "brand_or_video_context",
            "comment_text",
            "annotation_role",
            "final_human_label",
            "annotation_status",
            "source_file",
        ]
    ]


def codebook_lines() -> list[str]:
    examples = {
        "Positive": [
            "Aku cocok pakai serum ini, kulit terasa lebih halus.",
            "Rekomendasi banget buat yang cari pelembap ringan.",
            "Hasilnya pelan tapi bekas jerawatku mulai pudar.",
            "Produknya nyaman dan tidak bikin lengket.",
            "Wajib beli lagi kalau sudah habis.",
            "Suka sama teksturnya, cepat meresap.",
            "Akhirnya nemu sunscreen yang pas.",
            "Kulitku jadi terlihat lebih cerah setelah rutin pakai.",
            "Mantap, packaging aman dan isinya bagus.",
            "Aku percaya produk ini karena sejauh ini cocok.",
            "Bagus untuk kulitku yang mudah kering.",
            "Worth it dengan harga segitu.",
            "Temanku pakai dan hasilnya kelihatan baik.",
            "Semoga brand ini terus keluarin produk seperti ini.",
            "Ini penyelamat kulitku waktu lagi kusam.",
        ],
        "Neutral": [
            "Harganya berapa?",
            "Beli di mana ya kak?",
            "Ini dipakai pagi atau malam?",
            "Kandungan utamanya apa?",
            "Aku baru mau coba minggu depan.",
            "Ukuran botolnya berapa ml?",
            "Ada link produknya?",
            "Untuk umur 17 boleh tidak?",
            "Aku pakai merek lain sekarang.",
            "Ini varian yang mana?",
            "Kak spill urutan pemakaiannya.",
            "Belum coba, masih cari info.",
            "Tag teman yang tanya produk ini.",
            "Ini sama dengan yang di video sebelumnya?",
            "Pengiriman ke luar kota bisa?",
        ],
        "Negative": [
            "Aku tidak cocok, malah muncul jerawat.",
            "Baunya mengganggu dan bikin pusing.",
            "Sudah pakai lama tapi tidak ada perubahan.",
            "Kulitku terasa perih setelah pakai.",
            "Kecewa karena isinya bocor.",
            "Produk ini bikin wajahku makin berminyak.",
            "Aku kapok beli lagi.",
            "Teksturnya berat dan lengket.",
            "Klaimnya tidak sesuai di kulitku.",
            "Muka jadi gatal setelah pemakaian.",
            "Warnanya berubah dan aku ragu aman.",
            "Tidak worth it untuk harganya.",
            "Packaging rusak waktu sampai.",
            "Aku merasa efeknya buruk.",
            "Bukannya membaik, kulitku malah kusam.",
        ],
        "Uncertain": [
            "Hmm begitu ya.",
            "Yang ini gimana sih sebenarnya?",
            "Aduh kok bisa gitu.",
            "Aku kira beda.",
            "Mungkin iya mungkin tidak.",
            "Tergantung kulit masing-masing kali ya.",
            "Baru lihat, belum paham.",
            "Komentarnya bercanda tapi konteksnya kurang jelas.",
            "Wah menarik juga.",
            "Entah ini bagus atau tidak.",
            "Aku bingung maksud videonya.",
            "Kalimatnya terpotong dan ambigu.",
            "Nada komentarnya bisa sarkas, tapi tidak jelas.",
            "Tidak tahu ini pujian atau kritik.",
            "Butuh konteks video untuk menilai.",
        ],
        "No Text": [
            "",
            "   ",
            "[deleted]",
            "deleted",
            "<teks tidak tersedia>",
            "<komentar kosong>",
            "<hanya artefak teknis>",
            "<file tidak memuat isi komentar>",
            "<teks rusak tidak terbaca>",
            "<hanya whitespace>",
            "<baris tanpa comment text>",
            "<konten hilang saat ekspor>",
            "<placeholder kosong>",
            "<komentar tidak terscrape>",
            "<tidak ada teks yang dapat dinilai>",
        ],
    }
    lines = [
        "# SENTIMENT ANNOTATION CODEBOOK V4",
        "",
        "Positive: Komentar menunjukkan dukungan, kepuasan, rekomendasi, pengalaman baik, pujian, harapan positif, atau orientasi positif terhadap produk/konteks video.",
        "",
        "Neutral: Komentar bersifat informasional, pertanyaan tanpa evaluasi, pernyataan faktual, tagging, permintaan informasi, atau tidak menunjukkan evaluasi positif/negatif yang cukup.",
        "",
        "Negative: Komentar menunjukkan ketidakpuasan, kritik, keluhan, penolakan, pengalaman buruk, kekhawatiran negatif, atau evaluasi negatif terhadap produk/konteks.",
        "",
        "Uncertain: Teks tersedia tetapi makna sentimen tidak dapat ditentukan secara memadai, terlalu ambigu, sarkastik tanpa konteks cukup, atau konflik interpretasi tidak dapat diselesaikan.",
        "",
        "No Text: Komentar kosong, hanya whitespace, hanya artefak teknis, atau tidak mempunyai teks yang dapat dinilai.",
        "",
        "Aturan khusus:",
        "",
        "- Pertanyaan tidak otomatis Neutral; nilai orientasi evaluatifnya.",
        "- Promosi tidak otomatis Positive.",
        "- Emoji harus dibaca bersama teks.",
        "- Sarkasme harus menggunakan konteks yang tersedia.",
        "- Nama brand tidak menentukan label.",
        "- HCC tidak menentukan label.",
        "- Kata 'bagus' dalam negasi bukan Positive.",
        "- 'Belum coba' umumnya Neutral kecuali memiliki evaluasi lain.",
        "- 'Mau coba' dapat Neutral bila hanya menyatakan niat.",
        "- 'Wajib beli' dapat Positive bila menunjukkan rekomendasi.",
        "- 'Aman nggak?' umumnya Neutral/Uncertain tergantung konteks.",
        "- Label didasarkan pada isi komentar, bukan dugaan identitas atau motif akun.",
        "",
        "Contoh sintetis:",
        "",
    ]
    for label, items in examples.items():
        lines.append(f"## {label}")
        for idx, item in enumerate(items, start=1):
            lines.append(f"{idx}. {item}")
        lines.append("")
    return lines


def write_codebook() -> None:
    (OUT_DIR / "SENTIMENT_ANNOTATION_CODEBOOK_V4.md").write_text("\n".join(codebook_lines()) + "\n", encoding="utf-8")


def write_sheet(ws, df: pd.DataFrame, freeze: bool = True) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if freeze:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 18,
        "B": 22,
        "C": 22,
        "D": 22,
        "E": 24,
        "F": 70,
        "G": 16,
        "H": 18,
        "I": 14,
        "J": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_validations(ws, n_rows: int) -> None:
    if n_rows <= 0:
        return
    label_dv = DataValidation(type="list", formula1='"Negative,Neutral,Positive,Uncertain,No Text"', allow_blank=True)
    conf_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ctx_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(label_dv)
    ws.add_data_validation(conf_dv)
    ws.add_data_validation(ctx_dv)
    label_dv.add(f"G2:G{n_rows + 1}")
    conf_dv.add(f"H2:H{n_rows + 1}")
    ctx_dv.add(f"I2:I{n_rows + 1}")


def workbook(path: Path, sheets: dict[str, pd.DataFrame], readme_lines: list[str] | None = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    if readme_lines is not None:
        ws = wb.create_sheet("README")
        for line in readme_lines:
            ws.append([line])
        ws.column_dimensions["A"].width = 120
    for name, df in sheets.items():
        ws = wb.create_sheet(name[:31])
        write_sheet(ws, df)
        if set(ANNOTATOR_COLUMNS).issubset(df.columns):
            apply_validations(ws, len(df))
    wb.save(path)


def write_workbooks(master: pd.DataFrame, summary: pd.DataFrame) -> None:
    dev_a1 = annotator_view(master, "DEVELOPMENT_NEW_PENDING")
    dev_a2 = dev_a1.copy()
    dev_adj = adjudication_view(master, "DEVELOPMENT_NEW_PENDING")
    lock_a1 = annotator_view(master, "LOCKED_TEST_NEW_PENDING")
    lock_a2 = lock_a1.copy()
    lock_adj = adjudication_view(master, "LOCKED_TEST_NEW_PENDING")
    historical = historical_view(master)
    codebook_df = pd.DataFrame({"codebook": codebook_lines()})
    readme = [
        "MASTER_ANNOTATION_V4_READY_FOR_HUMAN_LABELING",
        "Isi hanya kolom human_label, confidence_annotation, needs_context, dan annotator_notes pada sheet annotator.",
        "Annotator tidak melihat prediksi model, probability, HCC, actor type, goal orientation, atau label annotator lain.",
    ]
    workbook(
        OUT_DIR / "human_annotation_master_v4.xlsx",
        {
            "LABEL_CODEBOOK": codebook_df,
            "DEVELOPMENT_ANNOTATOR_1": dev_a1,
            "DEVELOPMENT_ANNOTATOR_2": dev_a2,
            "DEVELOPMENT_ADJUDICATION": dev_adj,
            "LOCKED_TEST_ANNOTATOR_1": lock_a1,
            "LOCKED_TEST_ANNOTATOR_2": lock_a2,
            "LOCKED_TEST_ADJUDICATION": lock_adj,
            "HISTORICAL_FINAL_READONLY": historical,
            "AUDIT_SUMMARY": summary,
        },
        readme,
    )
    workbook(OUT_DIR / "sentiment_v4_development_annotator_1.xlsx", {"DEVELOPMENT_ANNOTATOR_1": dev_a1}, readme)
    workbook(OUT_DIR / "sentiment_v4_development_annotator_2.xlsx", {"DEVELOPMENT_ANNOTATOR_2": dev_a2}, readme)
    workbook(OUT_DIR / "sentiment_v4_development_adjudication.xlsx", {"DEVELOPMENT_ADJUDICATION": dev_adj}, readme)
    workbook(OUT_DIR / "sentiment_v4_locked_test_annotator_1.xlsx", {"LOCKED_TEST_ANNOTATOR_1": lock_a1}, readme)
    workbook(OUT_DIR / "sentiment_v4_locked_test_annotator_2.xlsx", {"LOCKED_TEST_ANNOTATOR_2": lock_a2}, readme)
    workbook(OUT_DIR / "sentiment_v4_locked_test_adjudication.xlsx", {"LOCKED_TEST_ADJUDICATION": lock_adj}, readme)


def target_plan() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"split": "historical_development_current", "target": "already_final", "negative": 147, "neutral": 286, "positive": 113, "total": 546},
            {"split": "development_pending_v4", "target": "1300 pending candidates", "negative": "", "neutral": "", "positive": "", "total": 1300},
            {"split": "locked_test_pending_v4", "target": "700 pending candidates", "negative": "", "neutral": "", "positive": "", "total": 700},
            {"split": "development_final_goal", "target": ">=1450 evaluable; ideal 1500-1700", "negative": ">=400", "neutral": ">=650", "positive": ">=400", "total": ">=1450"},
            {"split": "locked_test_final_goal", "target": ">=600 evaluable", "negative": ">=150", "neutral": ">=300", "positive": ">=150", "total": ">=600"},
        ]
    )


def readiness_report(master: pd.DataFrame, split_leakage: pd.DataFrame) -> pd.DataFrame:
    historical = master.loc[master["annotation_role"].eq("HISTORICAL_DEVELOPMENT_FINAL")]
    rows = [
        {"metric": "status", "value": "MASTER_ANNOTATION_V4_READY_FOR_HUMAN_LABELING", "passed": True, "notes": ""},
        {"metric": "historical_development_total", "value": len(historical), "passed": True, "notes": ""},
        {"metric": "historical_development_negative", "value": int(historical["final_human_label"].eq("Negative").sum()), "passed": True, "notes": ""},
        {"metric": "historical_development_neutral", "value": int(historical["final_human_label"].eq("Neutral").sum()), "passed": True, "notes": ""},
        {"metric": "historical_development_positive", "value": int(historical["final_human_label"].eq("Positive").sum()), "passed": True, "notes": ""},
        {"metric": "pending_development", "value": int(master["annotation_role"].eq("DEVELOPMENT_NEW_PENDING").sum()), "passed": int(master["annotation_role"].eq("DEVELOPMENT_NEW_PENDING").sum()) == 1300, "notes": ""},
        {"metric": "pending_locked_test", "value": int(master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING").sum()), "passed": int(master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING").sum()) == 700, "notes": ""},
        {"metric": "total_pending_manual_annotation", "value": int(master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]).sum()), "passed": int(master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]).sum()) == 2000, "notes": ""},
        {"metric": "total_master_rows", "value": len(master), "passed": True, "notes": ""},
        {"metric": "hard_duplicate_leakage", "value": int(split_leakage.loc[split_leakage["status"].eq("FAIL"), "observed"].sum()), "passed": not split_leakage["status"].eq("FAIL").any(), "notes": ""},
        {"metric": "pending_inj_count", "value": int(master.loc[master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]), "comment_id"].map(is_inj).sum()), "passed": int(master.loc[master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]), "comment_id"].map(is_inj).sum()) == 0, "notes": ""},
        {
            "metric": "pending_missing_comment_text",
            "value": int(master.loc[master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]), "comment_text"].map(no_text).sum()),
            "passed": True,
            "notes": "Reported for annotators; No Text is an allowed manual label and is not auto-filled.",
        },
        {"metric": "master_workbook", "value": rel(OUT_DIR / "human_annotation_master_v4.xlsx"), "passed": True, "notes": ""},
        {"metric": "codebook", "value": rel(OUT_DIR / "SENTIMENT_ANNOTATION_CODEBOOK_V4.md"), "passed": True, "notes": ""},
    ]
    return pd.DataFrame(rows)


def write_sampling_summary(master: pd.DataFrame) -> pd.DataFrame:
    summary = (
        master.loc[master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"])]
        .groupby(["annotation_role", "sampling_stratum"], dropna=False)
        .size()
        .rename("count")
        .reset_index()
    )
    summary.to_csv(OUT_DIR / "master_sampling_summary.csv", index=False, encoding="utf-8-sig")
    return summary


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    paths = discover_annotation_sources()
    inventory, checksums, schema, counts = inventory_sources(paths)
    inventory.to_csv(OUT_DIR / "source_annotation_inventory.csv", index=False, encoding="utf-8-sig")
    checksums.to_csv(OUT_DIR / "source_annotation_checksums.csv", index=False, encoding="utf-8-sig")
    schema.to_csv(OUT_DIR / "source_annotation_schema_audit.csv", index=False, encoding="utf-8-sig")
    counts.to_csv(OUT_DIR / "source_annotation_counts.csv", index=False, encoding="utf-8-sig")

    exact_map, near_map = external_cluster_maps()
    historical = historical_master_rows()
    pending, internal_dev, internal_locked, sampling_audit = build_pending_rows(historical, exact_map, near_map)
    historical = add_cluster_columns(historical, exact_map, near_map)
    priority = {
        "LOCKED_TEST_NEW_PENDING": 0,
        "LOCKED_TEST_NEW_FINAL": 0,
        "DEVELOPMENT_NEW_PENDING": 1,
        "DEVELOPMENT_NEW_FINAL": 1,
        "HISTORICAL_DEVELOPMENT_FINAL": 2,
        "LEGACY_TEST_PROVENANCE": 3,
        "EXCLUDED": 9,
    }
    combined = pd.concat([pending, historical], ignore_index=True, sort=False)
    combined["_priority"] = combined["annotation_role"].map(priority).fillna(99).astype(int)
    combined = combined.sort_values(["comment_id", "_priority", "source_file"]).drop(columns=["_priority"])
    master = finalize_master(combined)

    duplicate_resolution, split_leakage, comment_conflicts, text_conflicts, exclusion = conflict_and_leakage_audits(master)
    duplicate_resolution.to_csv(OUT_DIR / "master_duplicate_resolution.csv", index=False, encoding="utf-8-sig")
    split_leakage.to_csv(OUT_DIR / "master_split_leakage_audit.csv", index=False, encoding="utf-8-sig")
    comment_conflicts.to_csv(OUT_DIR / "master_comment_id_conflicts.csv", index=False, encoding="utf-8-sig")
    text_conflicts.to_csv(OUT_DIR / "master_text_cluster_conflicts.csv", index=False, encoding="utf-8-sig")
    exclusion.to_csv(OUT_DIR / "master_exclusion_audit.csv", index=False, encoding="utf-8-sig")
    internal_dev.to_csv(OUT_DIR / "development_sampling_internal_audit.csv", index=False, encoding="utf-8-sig")
    internal_locked.to_csv(OUT_DIR / "locked_test_sampling_internal_audit.csv", index=False, encoding="utf-8-sig")
    sampling_audit.to_csv(OUT_DIR / "master_sampling_internal_selection_audit.csv", index=False, encoding="utf-8-sig")
    target_plan().to_csv(OUT_DIR / "master_class_target_plan.csv", index=False, encoding="utf-8-sig")
    sampling_summary = write_sampling_summary(master)
    readiness = readiness_report(master, split_leakage)
    readiness.to_csv(OUT_DIR / "master_readiness_report.csv", index=False, encoding="utf-8-sig")
    write_codebook()
    write_workbooks(master, sampling_summary)
    master.to_csv(OUT_DIR / "human_annotation_master_v4.csv", index=False, encoding="utf-8-sig")

    manifest = {
        "status": "MASTER_ANNOTATION_V4_READY_FOR_HUMAN_LABELING",
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "rules": {
            "label_source": "human annotation only",
            "model_predictions_used_for_sampling_only": True,
            "model_predictions_visible_to_annotators": False,
            "training_run": False,
            "threshold_selection": False,
            "locked_test_evaluation": False,
            "full_inference": False,
        },
        "counts": {
            "historical_development_final": int(master["annotation_role"].eq("HISTORICAL_DEVELOPMENT_FINAL").sum()),
            "development_pending": int(master["annotation_role"].eq("DEVELOPMENT_NEW_PENDING").sum()),
            "locked_test_pending": int(master["annotation_role"].eq("LOCKED_TEST_NEW_PENDING").sum()),
            "total_pending": int(master["annotation_role"].isin(["DEVELOPMENT_NEW_PENDING", "LOCKED_TEST_NEW_PENDING"]).sum()),
            "total_master_rows": int(len(master)),
        },
        "workbooks": {
            "master": rel(OUT_DIR / "human_annotation_master_v4.xlsx"),
            "development_annotator_1": rel(OUT_DIR / "sentiment_v4_development_annotator_1.xlsx"),
            "development_annotator_2": rel(OUT_DIR / "sentiment_v4_development_annotator_2.xlsx"),
            "development_adjudication": rel(OUT_DIR / "sentiment_v4_development_adjudication.xlsx"),
            "locked_test_annotator_1": rel(OUT_DIR / "sentiment_v4_locked_test_annotator_1.xlsx"),
            "locked_test_annotator_2": rel(OUT_DIR / "sentiment_v4_locked_test_annotator_2.xlsx"),
            "locked_test_adjudication": rel(OUT_DIR / "sentiment_v4_locked_test_adjudication.xlsx"),
        },
        "source_inventory_rows": int(len(inventory)),
        "input_sha256": {
            rel(PREVIOUS_REGISTRY): sha256_file(PREVIOUS_REGISTRY),
            rel(ACTIVE_LEARNING_BLIND): sha256_file(ACTIVE_LEARNING_BLIND),
            rel(LOCKED_TEMPLATE): sha256_file(LOCKED_TEMPLATE),
            rel(OBS_V2): sha256_file(OBS_V2),
            rel(DATASET): sha256_file(DATASET),
        },
    }
    (OUT_DIR / "MASTER_ANNOTATION_V4_MANIFEST.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if not readiness["passed"].astype(bool).all():
        failed = readiness.loc[~readiness["passed"].astype(bool)]
        raise AssertionError(f"Master annotation readiness failed:\n{failed.to_string(index=False)}")
    print(json.dumps(manifest["counts"] | {"status": manifest["status"]}, indent=2))


if __name__ == "__main__":
    main()
